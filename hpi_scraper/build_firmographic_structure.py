from __future__ import annotations

import json
import os
import re
import shutil
import time
from pathlib import Path
from typing import Any
from urllib.parse import urlencode

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
BACKUP = ROOT / "Backup"
INPUT = ROOT / "input" / "compnys.txt"
OUT_ROOT = ROOT / "Firmographic"
HTTP_TIMEOUT = 45

DATA_COLUMNS = [
    "company_name",
    "legal_name",
    "website_domain",
    "linkedin_url",
    "employee_count",
    "headcount_growth_1yr",
    "revenue_or_revenue_band",
    "industry",
    "sub_industry_naics_sic",
    "hq_location",
    "number_of_sites_locations_sg_global",
    "founded_year",
    "ownership_type",
    "funding_total",
]

API_REPORT_COLUMNS = [
    "Tool Name",
    "Category",
    "API Available (Y/N)",
    "Authentication Type",
    "Credits Used",
    "Rate Limit",
    "Companies Processed",
    "Error Rate (%)",
    "Latency (sec)",
    "Status Code / Error Flag",
    "Records Retrieved",
    "Gated Fields Flag",
    "missing_info",
    "Raw Export Saved (Y/N)",
]


def safe_slug(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_")


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    if isinstance(value, list):
        return " | ".join(clean(v) for v in value if clean(v))
    text = str(value).strip()
    if text.lower() in {"none", "nan", "null", "not available", "not verified"}:
        return ""
    return text


def load_env(path: Path) -> dict[str, str]:
    env: dict[str, str] = {}
    if not path.exists():
        return env
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        env[key.strip()] = value.strip().strip('"').strip("'")
    return env


def money(amount: Any, currency: str = "") -> str:
    if amount in (None, ""):
        return ""
    try:
        amount_text = str(int(float(amount)))
    except (TypeError, ValueError):
        amount_text = str(amount)
    return f"{amount_text} {currency}".strip()


def apollo_record(raw: dict[str, Any], company_name: str) -> dict[str, str]:
    org = raw.get("organization") or {}
    naics = clean(org.get("naics_codes"))
    sic = clean(org.get("sic_codes"))
    sub_parts = []
    for source in (org.get("industries"), org.get("secondary_industries")):
        if isinstance(source, list):
            sub_parts.extend(clean(item) for item in source if clean(item))
        elif clean(source):
            sub_parts.append(clean(source))
    if naics:
        sub_parts.append(f"NAICS: {naics}")
    if sic:
        sub_parts.append(f"SIC: {sic}")
    deduped_sub_parts = list(dict.fromkeys(part for part in sub_parts if part))
    hq_parts = [org.get("city"), org.get("state"), org.get("postal_code"), org.get("country")]
    hq_location = clean(org.get("raw_address")) or ", ".join(clean(x) for x in hq_parts if clean(x))
    growth = org.get("organization_headcount_twelve_month_growth")
    growth_text = f"{float(growth) * 100:.2f}%" if isinstance(growth, (int, float)) else ""
    retail_locations = org.get("retail_location_count")
    location_count = ""
    if isinstance(retail_locations, (int, float)) and retail_locations > 0:
        location_count = clean(int(retail_locations))
    elif clean(retail_locations) and clean(retail_locations) != "0":
        location_count = clean(retail_locations)
    return {
        "company_name": company_name,
        "legal_name": clean(org.get("name")),
        "website_domain": clean(org.get("primary_domain")),
        "linkedin_url": clean(org.get("linkedin_url")),
        "employee_count": clean(org.get("estimated_num_employees")),
        "headcount_growth_1yr": growth_text,
        "revenue_or_revenue_band": clean(org.get("annual_revenue_printed") or org.get("organization_revenue_printed") or org.get("annual_revenue")),
        "industry": clean(org.get("industry")),
        "sub_industry_naics_sic": " | ".join(deduped_sub_parts),
        "hq_location": hq_location,
        "number_of_sites_locations_sg_global": location_count,
        "founded_year": clean(org.get("founded_year")),
        "ownership_type": "Public Company" if clean(org.get("publicly_traded_symbol") or org.get("publicly_traded_exchange")) else "",
        "funding_total": "",
    }


def extract_response_json(response: requests.Response) -> Any:
    try:
        return response.json()
    except ValueError:
        return {"raw_text": response.text}


def save_live_response(path: Path, request_data: dict[str, Any], response: requests.Response, latency_sec: float) -> Any:
    body = extract_response_json(response)
    export = {
        "request": request_data,
        "response": {
            "status_code": response.status_code,
            "headers": {k: v for k, v in response.headers.items()},
            "latency_sec": round(latency_sec, 3),
            "text": response.text if not isinstance(body, (dict, list)) else "",
        },
        "json": body,
    }
    write_json(path, export)
    return body


def live_apollo_call(company: pd.Series, api_key: str, raw_out: Path) -> tuple[dict[str, Any], dict[str, Any]]:
    url = "https://api.apollo.io/api/v1/organizations/enrich"
    params = {
        "domain": clean(company.get("domain")),
        "name": clean(company.get("company_name")),
        "website": f"https://{clean(company.get('domain'))}" if clean(company.get("domain")) else "",
        "linkedin_url": clean(company.get("linkedin_url")),
    }
    params = {k: v for k, v in params.items() if v}
    headers = {
        "accept": "application/json",
        "Cache-Control": "no-cache",
        "x-api-key": api_key,
    }
    started = time.perf_counter()
    response = requests.get(url, headers=headers, params=params, timeout=HTTP_TIMEOUT)
    latency = time.perf_counter() - started
    slug = safe_slug(company["company_name"])
    body = save_live_response(
        raw_out / f"{slug}.json",
        {"method": "GET", "url": url, "params": params},
        response,
        latency,
    )
    raw_for_parser = body if isinstance(body, dict) else {}
    trace = {
        "company_name": company["company_name"],
        "status_code": response.status_code,
        "latency_sec": round(latency, 3),
        "success": response.ok and bool(raw_for_parser.get("organization")),
        "records_retrieved": 1 if response.ok and raw_for_parser.get("organization") else 0,
        "error": "" if response.ok else response.text[:300],
        "rate_limit": "; ".join(
            f"{k}: {v}" for k, v in response.headers.items() if k.lower().startswith("x-rate-limit")
        ),
    }
    return raw_for_parser, trace


def live_coresignal_call(company: pd.Series, api_key: str, raw_out: Path) -> tuple[dict[str, Any], dict[str, Any]]:
    base_url = "https://api.coresignal.com/cdapi/v2/company_multi_source/enrich"
    website = f"https://{clean(company.get('domain'))}" if clean(company.get("domain")) else ""
    headers = {"accept": "application/json", "apikey": api_key}
    started = time.perf_counter()
    response = requests.get(base_url, headers=headers, params={"website": website}, timeout=HTTP_TIMEOUT)
    latency = time.perf_counter() - started
    slug = safe_slug(company["company_name"])
    body = save_live_response(
        raw_out / f"{slug}.json",
        {"method": "GET", "url": base_url, "params": {"website": website}},
        response,
        latency,
    )
    raw_for_parser = body if isinstance(body, dict) else {}
    trace = {
        "company_name": company["company_name"],
        "status_code": response.status_code,
        "latency_sec": round(latency, 3),
        "success": response.ok and isinstance(raw_for_parser, dict) and bool(raw_for_parser),
        "records_retrieved": 1 if response.ok and isinstance(raw_for_parser, dict) and bool(raw_for_parser) else 0,
        "error": "" if response.ok else response.text[:300],
        "rate_limit": "; ".join(
            f"{k}: {v}" for k, v in response.headers.items() if "rate" in k.lower()
        ),
    }
    return raw_for_parser, trace


def apollo_usage_stats(api_key: str, raw_out: Path) -> dict[str, Any]:
    url = "https://api.apollo.io/api/v1/usage_stats/api_usage_stats"
    headers = {"accept": "application/json", "Cache-Control": "no-cache", "x-api-key": api_key}
    started = time.perf_counter()
    response = requests.post(url, headers=headers, timeout=HTTP_TIMEOUT)
    latency = time.perf_counter() - started
    diagnostics_dir = raw_out / "info_and_diagnostics"
    diagnostics_dir.mkdir(parents=True, exist_ok=True)
    body = save_live_response(
        diagnostics_dir / "_apollo_usage_stats_live.json",
        {"method": "POST", "url": url},
        response,
        latency,
    )
    if not response.ok or not isinstance(body, dict):
        return {}
    return body.get('["api/v1/organizations", "enrich"]') or {}


def coresignal_revenue(raw: dict[str, Any]) -> str:
    annual = raw.get("revenue_annual")
    if isinstance(annual, dict):
        preferred = annual.get("source_5_annual_revenue")
        if isinstance(preferred, dict) and preferred.get("annual_revenue") is not None:
            return money(preferred.get("annual_revenue"), preferred.get("annual_revenue_currency", ""))
        for value in annual.values():
            if isinstance(value, dict) and value.get("annual_revenue") is not None:
                return money(value.get("annual_revenue"), value.get("annual_revenue_currency", ""))
    revenue_range = raw.get("revenue_annual_range")
    if isinstance(revenue_range, dict):
        for value in revenue_range.values():
            if isinstance(value, dict):
                frm = value.get("annual_revenue_range_from")
                to = value.get("annual_revenue_range_to")
                currency = value.get("annual_revenue_range_currency", "")
                if frm and to:
                    return f"{money(frm, currency)} - {money(to, currency)}"
                if frm:
                    return f"{money(frm, currency)}+"
    return ""


def funding_total(raw: dict[str, Any]) -> str:
    totals: dict[str, float] = {}
    for round_data in raw.get("funding_rounds") or []:
        if not isinstance(round_data, dict):
            continue
        amount = round_data.get("amount_raised")
        currency = round_data.get("amount_raised_currency") or "USD"
        if amount is None:
            continue
        try:
            totals[currency] = totals.get(currency, 0.0) + float(amount)
        except (TypeError, ValueError):
            continue
    return " | ".join(money(v, k) for k, v in sorted(totals.items()))


def coresignal_record(raw: dict[str, Any], company_name: str) -> dict[str, str]:
    emp_change = raw.get("employees_count_change") or {}
    growth = emp_change.get("change_yearly_percentage")
    growth_text = f"{float(growth):.2f}%" if isinstance(growth, (int, float)) else ""
    locations = raw.get("company_locations_full") or []
    primary_location = next((x.get("location_address") for x in locations if isinstance(x, dict) and x.get("is_primary")), "")
    if not primary_location:
        primary_location = raw.get("hq_full_address") or raw.get("hq_location") or ""
    naics = clean(raw.get("naics_codes"))
    sic = clean(raw.get("sic_codes"))
    sub_parts = []
    if clean(raw.get("categories_and_keywords")):
        sub_parts.append(clean(raw.get("categories_and_keywords")))
    if naics:
        sub_parts.append(f"NAICS: {naics}")
    if sic:
        sub_parts.append(f"SIC: {sic}")
    return {
        "company_name": company_name,
        "legal_name": clean(raw.get("company_legal_name")),
        "website_domain": clean(raw.get("website_domain")),
        "linkedin_url": clean(raw.get("canonical_linkedin_url") or raw.get("linkedin_url")),
        "employee_count": clean(raw.get("employees_count")),
        "headcount_growth_1yr": growth_text,
        "revenue_or_revenue_band": coresignal_revenue(raw),
        "industry": clean(raw.get("industry")),
        "sub_industry_naics_sic": " | ".join(sub_parts),
        "hq_location": clean(primary_location),
        "number_of_sites_locations_sg_global": clean(len(locations) if locations else ""),
        "founded_year": clean(raw.get("founded_year")),
        "ownership_type": clean(raw.get("ownership_status") or ("Public" if raw.get("is_public") else "")),
        "funding_total": funding_total(raw),
    }


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")


def format_workbook(path: Path) -> None:
    wb = load_workbook(path)
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            max_len = max((len(str(cell.value)) for cell in col[:200] if cell.value is not None), default=10)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 55)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    wb.save(path)


def write_xlsx(path: Path, df: pd.DataFrame, sheet_name: str) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    format_workbook(path)


def missing_fields(rows: list[dict[str, str]]) -> list[str]:
    gated = []
    fields = [c for c in DATA_COLUMNS if c != "company_name"]
    for field in fields:
        if any(not clean(row.get(field)) for row in rows):
            gated.append(field)
    return gated


def build_tool(tool: str, companies: pd.DataFrame) -> None:
    tool_dir = OUT_ROOT / tool
    raw_out = tool_dir / "raw_exports"
    raw_out.mkdir(parents=True, exist_ok=True)
    data_rows: list[dict[str, str]] = []
    api_records: list[dict[str, str]] = []

    env = load_env(ROOT / ".env")

    if tool == "apollo":
        parser = apollo_record
        tool_name = "Apollo Organization Enrichment API"
        api_key = env.get("APOLLO_API_KEY", "")
        live_call = live_apollo_call
    else:
        parser = coresignal_record
        tool_name = "Coresignal Multi-source Company Enrichment API"
        api_key = env.get("CORESIGNAL_API_KEY", "")
        live_call = live_coresignal_call

    for _, company in companies.iterrows():
        if not api_key:
            continue
        try:
            raw, trace = live_call(company, api_key, raw_out)
        except requests.RequestException as exc:
            raw = {}
            trace = {
                "company_name": company["company_name"],
                "status_code": "REQUEST_ERROR",
                "latency_sec": 0,
                "success": False,
                "records_retrieved": 0,
                "error": str(exc),
                "rate_limit": "",
            }
        api_records.append(trace)
        if not trace["success"]:
            continue
        data_rows.append(parser(raw, company["company_name"]))

    gated = missing_fields(data_rows)
    statuses: dict[str, int] = {}
    for trace in api_records:
        key = str(trace["status_code"])
        statuses[key] = statuses.get(key, 0) + 1
    status_flag = "; ".join(f"{code}:{count}" for code, count in sorted(statuses.items())) or "NO_CALLS"
    success_count = sum(1 for trace in api_records if trace["success"])
    error_count = len(api_records) - success_count
    error_rate = f"{(error_count / len(api_records) * 100):.2f}" if api_records else "100.00"
    avg_latency = sum(float(trace["latency_sec"]) for trace in api_records) / len(api_records) if api_records else 0
    records = sum(int(trace["records_retrieved"]) for trace in api_records)
    rate_limits = [trace["rate_limit"] for trace in api_records if trace.get("rate_limit")]
    if tool == "apollo":
        usage = apollo_usage_stats(api_key, raw_out) if api_key else {}
        day = usage.get("day") if isinstance(usage, dict) else {}
        hour = usage.get("hour") if isinstance(usage, dict) else {}
        minute = usage.get("minute") if isinstance(usage, dict) else {}
        if day:
            credits = (
                f"{day.get('consumed', '')} organization enrich API units consumed today; "
                f"{hour.get('consumed', '')} consumed this hour"
            )
            rate_limit = (
                f"organization enrich day {day.get('limit', '')} consumed {day.get('consumed', '')} "
                f"left {day.get('left_over', '')}; hour {hour.get('limit', '')} consumed {hour.get('consumed', '')} "
                f"left {hour.get('left_over', '')}; minute {minute.get('limit', '')} left {minute.get('left_over', '')}"
            )
        else:
            credits = "Live org enrich calls made; usage-stats endpoint unavailable"
            rate_limit = rate_limits[0] if rate_limits else "Not returned in response headers"
        auth = "API Key in x-api-key header"
    else:
        credits = f"Estimated {success_count * 2} Collect credits ({success_count} successful enrich x 2)"
        rate_limit = rate_limits[0] if rate_limits else "Multi-source Company API documented enrichment rate limit: 18 req/sec."
        auth = "API Key in apikey header"
    api_available = "Y" if api_key else "N"

    summary_records = [
        {
            "Tool Name": tool_name,
            "Category": "Firmographic",
            "API Available (Y/N)": api_available,
            "Authentication Type": auth,
            "Credits Used": credits,
            "Rate Limit": rate_limit,
            "Companies Processed": str(len(companies)),
            "Error Rate (%)": error_rate,
            "Latency (sec)": f"{avg_latency:.3f} avg",
            "Status Code / Error Flag": status_flag,
            "Records Retrieved": str(records),
            "Gated Fields Flag": "Y - " + "; ".join(gated) if gated else "N",
            "missing_info": (
                "Apollo enrichment does not return funding_total for these public companies. "
                "Apollo also returns retail_location_count, not a full SG/global site-count field; "
                "zero retail counts are treated as unavailable, not as real zero locations."
                if tool == "apollo"
                else ""
            ),
            "Raw Export Saved (Y/N)": "Y",
        }
    ]

    data_json = tool_dir / f"{tool}_data.json"
    api_json = tool_dir / f"{tool}_api_records.json"
    write_json(data_json, data_rows)
    write_json(api_json, summary_records)
    write_json(tool_dir / f"{tool}_api_call_detail.json", api_records)

    data_df = pd.DataFrame(data_rows, columns=DATA_COLUMNS)
    api_df = pd.DataFrame(summary_records, columns=API_REPORT_COLUMNS)
    write_xlsx(tool_dir / f"{tool}_data.xlsx", data_df, f"{tool}_data")
    write_xlsx(tool_dir / f"{tool}_api_report.xlsx", api_df, f"{tool}_api_report")


def main() -> None:
    companies = pd.read_csv(INPUT, dtype=str).fillna("")
    # Remove old generated Firmographic output only; Backup remains untouched.
    if OUT_ROOT.exists():
        shutil.rmtree(OUT_ROOT)
    for tool in ("apollo", "coresignal"):
        build_tool(tool, companies)
    print(f"Wrote {OUT_ROOT}")
    for path in sorted(OUT_ROOT.glob("*/*")):
        if path.is_file():
            print(path.relative_to(ROOT))


if __name__ == "__main__":
    main()
