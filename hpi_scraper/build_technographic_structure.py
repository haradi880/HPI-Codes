from __future__ import annotations

import json
import re
import shutil
import time
from pathlib import Path
from typing import Any

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
INPUT = ROOT / "input" / "compnys.txt"
OUT_ROOT = ROOT / "Technographic"
HTTP_TIMEOUT = 60

DATA_COLUMNS = [
    "company_name",
    "website_domain",
    "detected_technologies_full_list",
    "hardware_device_endpoint_signals",
    "print_mps_collaboration_uc_stack",
    "it_spend_estimate",
    "detection_source",
    "last_seen_freshness_date_per_technology",
]

IT_SPEND_UNAVAILABLE = ""

API_REPORT_COLUMNS = [
    "Tool Name",
    "Category",
    "API Available (Y/N)",
    "Authentication Type",
    "Credits Used",
    "Rate Limit",
    "Companies Processed",
    "Error Rate (%)",
    "Latency (sec)",
    "Status Code / Error Flag",
    "Records Retrieved",
    "Gated Fields Flag",
    "missing_info",
    "Raw Export Saved (Y/N)",
]

HARDWARE_TERMS = {
    "endpoint",
    "device",
    "hardware",
    "printer",
    "print",
    "mps",
    "laptop",
    "desktop",
    "mobile",
    "android",
    "ios",
    "iphone",
    "ipad",
    "windows",
    "macos",
    "jamf",
    "intune",
    "cisco",
    "router",
    "switch",
    "firewall",
    "fortinet",
    "palo alto",
    "vmware",
}

PRINT_UC_TERMS = {
    "print",
    "printer",
    "mps",
    "managed print",
    "xerox",
    "ricoh",
    "canon",
    "hp",
    "lexmark",
    "collaboration",
    "uc",
    "unified communication",
    "teams",
    "microsoft teams",
    "zoom",
    "webex",
    "slack",
    "avaya",
    "voip",
    "pbx",
    "ringcentral",
}


def safe_slug(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_")


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    if isinstance(value, list):
        return " | ".join(clean(v) for v in value if clean(v))
    text = str(value).strip()
    if text.lower() in {"none", "nan", "null", "not available", "not verified"}:
        return ""
    return text


def dedupe(values: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for value in values:
        text = clean(value)
        key = text.lower()
        if text and key not in seen:
            seen.add(key)
            out.append(text)
    return out


def load_env(path: Path) -> dict[str, str]:
    env: dict[str, str] = {}
    if not path.exists():
        return env
    for line in path.read_text(encoding="utf-8").splitlines():
        if "=" not in line or line.strip().startswith("#"):
            continue
        key, value = line.split("=", 1)
        env[key.strip()] = value.strip().strip('"').strip("'")
    return env


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")


def format_workbook(path: Path) -> None:
    wb = load_workbook(path)
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column in ws.columns:
            letter = get_column_letter(column[0].column)
            max_len = max((len(str(cell.value)) for cell in column if cell.value is not None), default=10)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 70)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    wb.save(path)


def write_xlsx(path: Path, rows: list[dict[str, str]], columns: list[str], sheet_name: str) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        pd.DataFrame(rows, columns=columns).to_excel(writer, sheet_name=sheet_name, index=False)
    format_workbook(path)


def response_json(response: requests.Response) -> Any:
    try:
        return response.json()
    except ValueError:
        return {"raw_text": response.text}


def save_response(path: Path, request_data: dict[str, Any], response: requests.Response, latency_sec: float) -> Any:
    body = response_json(response)
    write_json(
        path,
        {
            "request": request_data,
            "response": {
                "status_code": response.status_code,
                "headers": dict(response.headers),
                "latency_sec": round(latency_sec, 3),
                "text": response.text if not isinstance(body, (dict, list)) else "",
            },
            "json": body,
        },
    )
    return body


def term_filter(technologies: list[dict[str, str]], terms: set[str]) -> list[str]:
    found: list[str] = []
    for tech in technologies:
        haystack = " ".join(
            [
                tech.get("name", ""),
                tech.get("category", ""),
                tech.get("parent_category", ""),
            ]
        ).lower()
        if any(term in haystack for term in terms):
            found.append(tech.get("name", ""))
    return dedupe(found)


def theirstack_balance(api_key: str, raw_dir: Path, label: str) -> dict[str, Any]:
    url = "https://api.theirstack.com/v0/billing/credit-balance"
    headers = {"Authorization": f"Bearer {api_key}", "accept": "application/json"}
    started = time.perf_counter()
    response = requests.get(url, headers=headers, timeout=HTTP_TIMEOUT)
    latency = time.perf_counter() - started
    body = save_response(
        raw_dir / "info_and_diagnostics" / f"_theirstack_credit_balance_{label}.json",
        {"method": "GET", "url": url},
        response,
        latency,
    )
    return body if isinstance(body, dict) else {}


def live_theirstack(company: pd.Series, api_key: str, raw_dir: Path) -> tuple[dict[str, Any], dict[str, Any]]:
    url = "https://api.theirstack.com/v1/companies/technologies"
    payload = {
        "company_domain": clean(company["domain"]),
        "limit": 500,
        "page": 0,
    }
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
        "accept": "application/json",
    }
    started = time.perf_counter()
    response = requests.post(url, headers=headers, json=payload, timeout=HTTP_TIMEOUT)
    latency = time.perf_counter() - started
    body = save_response(
        raw_dir / f"{safe_slug(company['company_name'])}.json",
        {"method": "POST", "url": url, "payload": payload},
        response,
        latency,
    )
    records = len(body.get("data") or []) if isinstance(body, dict) else 0
    return (body if isinstance(body, dict) else {}), {
        "company_name": company["company_name"],
        "status_code": response.status_code,
        "latency_sec": round(latency, 3),
        "success": response.ok,
        "records_retrieved": records,
        "error": "" if response.ok else response.text[:500],
        "rate_limit": "; ".join(f"{k}: {v}" for k, v in response.headers.items() if "rate" in k.lower()),
    }


def live_coresignal(company: pd.Series, api_key: str, raw_dir: Path) -> tuple[dict[str, Any], dict[str, Any]]:
    url = "https://api.coresignal.com/cdapi/v2/company_multi_source/enrich"
    params = {"website": f"https://{clean(company['domain'])}"}
    headers = {"accept": "application/json", "apikey": api_key}
    started = time.perf_counter()
    response = requests.get(url, headers=headers, params=params, timeout=HTTP_TIMEOUT)
    latency = time.perf_counter() - started
    body = save_response(
        raw_dir / f"{safe_slug(company['company_name'])}.json",
        {"method": "GET", "url": url, "params": params},
        response,
        latency,
    )
    records = 1 if response.ok and isinstance(body, dict) and body else 0
    return (body if isinstance(body, dict) else {}), {
        "company_name": company["company_name"],
        "status_code": response.status_code,
        "latency_sec": round(latency, 3),
        "success": response.ok,
        "records_retrieved": records,
        "error": "" if response.ok else response.text[:500],
        "rate_limit": "; ".join(f"{k}: {v}" for k, v in response.headers.items() if "rate" in k.lower()),
    }


def parse_theirstack(raw: dict[str, Any], company: pd.Series) -> dict[str, str]:
    technologies: list[dict[str, str]] = []
    freshness: list[str] = []
    for item in raw.get("data") or []:
        tech = item.get("technology") or {}
        name = clean(tech.get("name"))
        if not name:
            continue
        technologies.append(
            {
                "name": name,
                "category": clean(tech.get("category")),
                "parent_category": clean(tech.get("parent_category")),
            }
        )
        last_seen = clean(item.get("last_date_found"))
        first_seen = clean(item.get("first_date_found"))
        confidence = clean(item.get("confidence"))
        jobs = clean(item.get("jobs"))
        freshness.append(f"{name}: last_seen={last_seen or 'not_returned'}, first_seen={first_seen or 'not_returned'}, confidence={confidence}, jobs={jobs}")
    tech_names = dedupe([t["name"] for t in technologies])
    return {
        "company_name": clean(company["company_name"]),
        "website_domain": clean(company["domain"]),
        "detected_technologies_full_list": "; ".join(tech_names),
        "hardware_device_endpoint_signals": "; ".join(term_filter(technologies, HARDWARE_TERMS)),
        "print_mps_collaboration_uc_stack": "; ".join(term_filter(technologies, PRINT_UC_TERMS)),
        "it_spend_estimate": IT_SPEND_UNAVAILABLE,
        "detection_source": "TheirStack technographics API; detected from job-posting technology mentions",
        "last_seen_freshness_date_per_technology": " | ".join(freshness),
    }


def parse_coresignal(raw: dict[str, Any], company: pd.Series) -> dict[str, str]:
    tech_items = raw.get("technologies_used") or raw.get("current_technologies") or []
    technologies: list[dict[str, str]] = []
    freshness: list[str] = []
    if isinstance(tech_items, list):
        for item in tech_items:
            if isinstance(item, dict):
                name = clean(item.get("technology") or item.get("name"))
                technologies.append(
                    {
                        "name": name,
                        "category": clean(item.get("category")),
                        "parent_category": "",
                    }
                )
                if name:
                    first_seen = clean(item.get("first_verified_at") or item.get("first_seen_at"))
                    last_seen = clean(item.get("last_verified_at") or item.get("last_seen_at"))
                    freshness.append(f"{name}: last_seen={last_seen or 'not_returned'}, first_seen={first_seen or 'not_returned'}")
            elif clean(item):
                technologies.append({"name": clean(item), "category": "", "parent_category": ""})
    if not technologies:
        technologies = [{"name": name, "category": "", "parent_category": ""} for name in raw.get("technology_names") or []]
    tech_names = dedupe([t["name"] for t in technologies])
    return {
        "company_name": clean(company["company_name"]),
        "website_domain": clean(company["domain"]),
        "detected_technologies_full_list": "; ".join(tech_names),
        "hardware_device_endpoint_signals": "; ".join(term_filter(technologies, HARDWARE_TERMS)),
        "print_mps_collaboration_uc_stack": "; ".join(term_filter(technologies, PRINT_UC_TERMS)),
        "it_spend_estimate": IT_SPEND_UNAVAILABLE,
        "detection_source": "Coresignal Multi-source Company Enrichment API; provider returned technologies_used with verification dates",
        "last_seen_freshness_date_per_technology": " | ".join(freshness),
    }


def blank_row(company: pd.Series, source_note: str) -> dict[str, str]:
    return {
        "company_name": clean(company["company_name"]),
        "website_domain": clean(company["domain"]),
        "detected_technologies_full_list": "",
        "hardware_device_endpoint_signals": "",
        "print_mps_collaboration_uc_stack": "",
        "it_spend_estimate": IT_SPEND_UNAVAILABLE,
        "detection_source": source_note,
        "last_seen_freshness_date_per_technology": "",
    }


def missing_fields(rows: list[dict[str, str]]) -> list[str]:
    gated: list[str] = []
    for field in DATA_COLUMNS:
        if field in {"company_name", "website_domain"}:
            continue
        if any(not clean(row.get(field)) for row in rows):
            gated.append(field)
    return gated


def summary_record(tool: str, traces: list[dict[str, Any]], rows: list[dict[str, str]], credits: str, auth: str, missing_info: str) -> dict[str, str]:
    statuses: dict[str, int] = {}
    for trace in traces:
        key = str(trace["status_code"])
        statuses[key] = statuses.get(key, 0) + 1
    successes = sum(1 for trace in traces if trace.get("success"))
    error_rate = ((len(traces) - successes) / len(traces) * 100) if traces else 100
    latency = sum(float(trace.get("latency_sec") or 0) for trace in traces) / len(traces) if traces else 0
    rate_limits = [clean(trace.get("rate_limit")) for trace in traces if clean(trace.get("rate_limit"))]
    return {
        "Tool Name": tool,
        "Category": "Technographic",
        "API Available (Y/N)": "Y",
        "Authentication Type": auth,
        "Credits Used": credits,
        "Rate Limit": rate_limits[0] if rate_limits else "Not returned in response headers",
        "Companies Processed": str(len(traces)),
        "Error Rate (%)": f"{error_rate:.2f}",
        "Latency (sec)": f"{latency:.3f} avg",
        "Status Code / Error Flag": "; ".join(f"{code}:{count}" for code, count in sorted(statuses.items())),
        "Records Retrieved": str(sum(int(trace.get("records_retrieved") or 0) for trace in traces)),
        "Gated Fields Flag": "Y - " + "; ".join(missing_fields(rows)) if missing_fields(rows) else "N",
        "missing_info": missing_info,
        "Raw Export Saved (Y/N)": "Y",
    }


def build_tool(tool: str, companies: pd.DataFrame, env: dict[str, str]) -> None:
    tool_dir = OUT_ROOT / tool
    raw_dir = tool_dir / "raw_exports"
    raw_dir.mkdir(parents=True, exist_ok=True)
    rows: list[dict[str, str]] = []
    traces: list[dict[str, Any]] = []

    if tool == "theirstack":
        api_key = env.get("THEIRSTACK_API_KEY", "")
        balance_before = theirstack_balance(api_key, raw_dir, "before") if api_key else {}
        for _, company in companies.iterrows():
            raw, trace = live_theirstack(company, api_key, raw_dir)
            traces.append(trace)
            if trace["success"]:
                rows.append(parse_theirstack(raw, company))
            else:
                rows.append(blank_row(company, f"TheirStack API returned {trace['status_code']}: {trace.get('error', '')[:250]}"))
        balance_after = theirstack_balance(api_key, raw_dir, "after") if api_key else {}
        before_used = balance_before.get("used_api_credits")
        after_used = balance_after.get("used_api_credits")
        if isinstance(before_used, int) and isinstance(after_used, int):
            credits = f"{after_used - before_used} API credits used in this run; docs state 3 credits per company lookup"
        else:
            successful_responses = sum(1 for trace in traces if int(trace.get("records_retrieved") or 0) > 0)
            credits = f"Estimated {successful_responses * 3} API credits; docs state 3 credits per company lookup with response"
        report = summary_record(
            "TheirStack Technographics API",
            traces,
            rows,
            credits,
            "Bearer token in Authorization header",
            "IT spend estimate is not returned by TheirStack technographics. Hardware/print/UC fields are extracted by category/name matching from returned technology objects.",
        )
    else:
        api_key = env.get("CORESIGNAL_API_KEY", "")
        for _, company in companies.iterrows():
            raw, trace = live_coresignal(company, api_key, raw_dir)
            traces.append(trace)
            if trace["success"]:
                rows.append(parse_coresignal(raw, company))
            else:
                rows.append(blank_row(company, f"Coresignal API returned {trace['status_code']}: {trace.get('error', '')[:250]}"))
        report = summary_record(
            "Coresignal Multi-source Company Enrichment API",
            traces,
            rows,
            f"Estimated {len(rows) * 2} Collect credits ({len(rows)} successful enrich x 2)",
            "API Key in apikey header",
            "Coresignal enrich returns technologies_used with first/last verification dates, but does not return IT spend estimate in this response.",
        )

    write_json(tool_dir / f"{tool}_data.json", rows)
    write_json(tool_dir / f"{tool}_api_records.json", [report])
    write_json(tool_dir / f"{tool}_api_call_detail.json", traces)
    write_xlsx(tool_dir / f"{tool}_data.xlsx", rows, DATA_COLUMNS, f"{tool}_data")
    write_xlsx(tool_dir / f"{tool}_api_report.xlsx", [report], API_REPORT_COLUMNS, f"{tool}_api_report")


def main() -> None:
    env = load_env(ROOT / ".env")
    companies = pd.read_csv(INPUT, dtype=str).fillna("")
    if OUT_ROOT.exists():
        shutil.rmtree(OUT_ROOT)
    for tool in ("theirstack", "coresignal"):
        build_tool(tool, companies, env)
    print(f"Wrote {OUT_ROOT}")
    for path in sorted(OUT_ROOT.glob("*/*")):
        if path.is_file():
            print(path.relative_to(ROOT))


if __name__ == "__main__":
    main()
