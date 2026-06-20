from __future__ import annotations

import json
import re
import shutil
import time
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any
from urllib.parse import quote_plus

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent

INPUT = ROOT / "input" / "compnys.txt"
OUT_ROOT = ROOT / "NewsAnnouncements"

HTTP_TIMEOUT = 45
MAX_RESULTS_PER_COMPANY_PER_TOOL = 5
LOOKBACK_DAYS = 365

DATA_COLUMNS = [
    "company_name",
    "website_domain",
    "company_linkedin_url",
    "news_announcements",
    "event_headline",
    "event_url",
    "event_date",
    "event_type",
    "source_publisher",
    "relevance_confidence",
    "coverage_depth_events_per_account_last_12mo",
]

API_REPORT_COLUMNS = [
    "Tool Name",
    "Category",
    "API Available (Y/N)",
    "Authentication Type",
    "Credits Used",
    "Rate Limit",
    "Companies Processed",
    "Error Rate (%)",
    "Latency (sec)",
    "Status Code / Error Flag",
    "Records Retrieved",
    "Gated Fields Flag",
    "missing_info",
    "Raw Export Saved (Y/N)",
]


EVENT_PATTERNS = [
    ("funding", re.compile(r"\b(funding|fundraise|raised|investment|series [a-z]|grant|financing|capital raise)\b", re.I)),
    ("expansion", re.compile(r"\b(expansion|expand|opens|new office|new facility|factory|plant|regional hub|data center|datacenter|market entry)\b", re.I)),
    ("launch", re.compile(r"\b(launch|unveil|release|introduce|roll out|debut|announces new product|new service)\b", re.I)),
    ("leadership", re.compile(r"\b(appoint|appointed|ceo|cfo|cto|cio|chairman|president|leadership|executive|resigns|steps down)\b", re.I)),
    ("M&A", re.compile(r"\b(acquire|acquisition|merger|merge|m&a|takeover|buyout|stake|joint venture|divest|sell unit)\b", re.I)),
]


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")


def safe_slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value).lower()).strip("_") or "item"


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    if isinstance(value, list):
        return " | ".join(clean(v) for v in value if clean(v))
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)[:300]
    text = str(value).strip()
    if text.lower() in {"none", "nan", "null", "not available", "n/a"}:
        return ""
    return text


def load_env(path: Path) -> dict[str, str]:
    env: dict[str, str] = {}
    if not path.exists():
        return env

    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        env[key.strip()] = value.strip().strip('"').strip("'")

    return env


def read_companies() -> pd.DataFrame:
    if not INPUT.exists():
        raise FileNotFoundError(f"Input file not found: {INPUT}")

    df = pd.read_csv(INPUT, dtype=str).fillna("")
    df.columns = [c.strip() for c in df.columns]

    required = ["company_name", "domain", "linkedin_url"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Missing required input columns: {missing}")

    df["company_name"] = df["company_name"].map(clean)
    df["domain"] = df["domain"].map(lambda x: clean(x).replace("https://", "").replace("http://", "").strip("/"))
    df["linkedin_url"] = df["linkedin_url"].map(clean)

    return df[(df["company_name"] != "") & (df["domain"] != "")].copy()


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")


def format_workbook(path: Path) -> None:
    wb = load_workbook(path)
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            max_len = max((len(str(cell.value)) for cell in col[:200] if cell.value is not None), default=10)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 65)

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(path)


def write_xlsx(path: Path, df: pd.DataFrame, sheet_name: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    format_workbook(path)


def response_json(response: requests.Response) -> Any:
    try:
        return response.json()
    except ValueError:
        return {"raw_text": response.text}


def save_response(path: Path, request_data: dict[str, Any], response: requests.Response, latency_sec: float) -> Any:
    body = response_json(response)

    write_json(
        path,
        {
            "request": request_data,
            "response": {
                "status_code": response.status_code,
                "headers": dict(response.headers),
                "latency_sec": round(latency_sec, 3),
                "text": response.text if not isinstance(body, (dict, list)) else "",
            },
            "json": body,
        },
    )

    return body


def save_text_response(path: Path, request_data: dict[str, Any], response: requests.Response, latency_sec: float) -> str:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(
            {
                "request": request_data,
                "response": {
                    "status_code": response.status_code,
                    "headers": dict(response.headers),
                    "latency_sec": round(latency_sec, 3),
                    "text": response.text,
                },
            },
            indent=2,
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    return response.text


def request_json_api(
    method: str,
    url: str,
    *,
    headers: dict[str, str],
    raw_path: Path,
    params: dict[str, Any] | None = None,
    payload: dict[str, Any] | None = None,
) -> tuple[Any, dict[str, Any]]:

    started = time.perf_counter()

    try:
        response = requests.request(
            method,
            url,
            headers=headers,
            params=params,
            json=payload,
            timeout=HTTP_TIMEOUT,
        )
        latency = time.perf_counter() - started
        body = save_response(
            raw_path,
            {
                "method": method,
                "url": url,
                "params": params or {},
                "payload": payload or {},
            },
            response,
            latency,
        )

        trace = {
            "status_code": response.status_code,
            "latency_sec": round(latency, 3),
            "success": response.ok,
            "records_retrieved": count_records(body),
            "error": "" if response.ok else response.text[:500],
            "rate_limit": "; ".join(
                f"{k}: {v}"
                for k, v in response.headers.items()
                if "rate" in k.lower() or "limit" in k.lower() or "credit" in k.lower()
            ),
        }

        return body, trace

    except Exception as exc:
        latency = time.perf_counter() - started
        body = {"error": str(exc)}

        write_json(
            raw_path,
            {
                "request": {"method": method, "url": url, "params": params or {}, "payload": payload or {}},
                "response": {"status_code": "EXCEPTION", "latency_sec": round(latency, 3)},
                "json": body,
            },
        )

        return body, {
            "status_code": "EXCEPTION",
            "latency_sec": round(latency, 3),
            "success": False,
            "records_retrieved": 0,
            "error": str(exc),
            "rate_limit": "",
        }


def request_text_api(
    method: str,
    url: str,
    *,
    raw_path: Path,
    headers: dict[str, str] | None = None,
    params: dict[str, Any] | None = None,
) -> tuple[str, dict[str, Any]]:

    started = time.perf_counter()

    try:
        response = requests.request(
            method,
            url,
            headers=headers or {},
            params=params,
            timeout=HTTP_TIMEOUT,
        )
        latency = time.perf_counter() - started

        text = save_text_response(
            raw_path,
            {"method": method, "url": url, "params": params or {}},
            response,
            latency,
        )

        record_count = text.count("<item>")

        return text, {
            "status_code": response.status_code,
            "latency_sec": round(latency, 3),
            "success": response.ok,
            "records_retrieved": record_count,
            "error": "" if response.ok else response.text[:500],
            "rate_limit": "",
        }

    except Exception as exc:
        latency = time.perf_counter() - started
        write_json(
            raw_path,
            {
                "request": {"method": method, "url": url, "params": params or {}},
                "response": {"status_code": "EXCEPTION", "latency_sec": round(latency, 3)},
                "text": "",
                "error": str(exc),
            },
        )

        return "", {
            "status_code": "EXCEPTION",
            "latency_sec": round(latency, 3),
            "success": False,
            "records_retrieved": 0,
            "error": str(exc),
            "rate_limit": "",
        }


def count_records(body: Any) -> int:
    if isinstance(body, dict):
        for key in ["results", "articles", "data", "items"]:
            if isinstance(body.get(key), list):
                return len(body[key])
    if isinstance(body, list):
        return len(body)
    return 0


def build_query(company_name: str, domain: str) -> str:
    return (
        f'"{company_name}" OR "{domain}" '
        f'(funding OR expansion OR launch OR leadership OR acquisition OR merger OR partnership OR appointed OR opens)'
    )


def infer_event_type(text: str) -> str:
    for label, pattern in EVENT_PATTERNS:
        if pattern.search(text):
            return label
    return "other"


def parse_date(value: Any) -> str:
    text = clean(value)
    if not text:
        return ""

    formats = [
        "%Y-%m-%dT%H:%M:%SZ",
        "%Y-%m-%dT%H:%M:%S.%fZ",
        "%Y-%m-%d",
        "%a, %d %b %Y %H:%M:%S %Z",
        "%a, %d %b %Y %H:%M:%S %z",
    ]

    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue

    match = re.search(r"\d{4}-\d{2}-\d{2}", text)
    if match:
        return match.group(0)

    return text[:30]


def is_last_12_months(date_text: str) -> bool:
    parsed = parse_date(date_text)
    if not re.match(r"^\d{4}-\d{2}-\d{2}$", parsed):
        return True

    event_date = datetime.strptime(parsed, "%Y-%m-%d").date()
    cutoff = (datetime.now(timezone.utc) - timedelta(days=LOOKBACK_DAYS)).date()

    return event_date >= cutoff


def confidence(company_name: str, domain: str, title: str, url: str, summary: str) -> str:
    blob = f"{title} {url} {summary}".lower()
    score = 0

    if company_name.lower() in blob:
        score += 50

    domain_root = domain.split(".")[0].lower()
    if domain_root and domain_root in blob:
        score += 20

    if infer_event_type(blob) != "other":
        score += 20

    if url:
        score += 10

    if score >= 80:
        return "High"
    if score >= 50:
        return "Medium"
    return "Low"


def normalize_event(
    *,
    company_name: str,
    domain: str,
    company_linkedin_url: str,
    title: str,
    url: str,
    date: str,
    publisher: str,
    summary: str,
) -> dict[str, str]:

    headline = clean(title)
    event_url = clean(url)
    event_date = parse_date(date)
    source = clean(publisher)
    blob = f"{headline} {summary}"

    return {
        "company_name": company_name,
        "website_domain": domain,
        "company_linkedin_url": company_linkedin_url,
        "news_announcements": headline,
        "event_headline": headline,
        "event_url": event_url,
        "event_date": event_date,
        "event_type": infer_event_type(blob),
        "source_publisher": source,
        "relevance_confidence": confidence(company_name, domain, headline, event_url, summary),
        "coverage_depth_events_per_account_last_12mo": "",
    }


def dedupe_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    seen: set[str] = set()
    out: list[dict[str, str]] = []

    for row in rows:
        key = clean(row.get("event_url")).lower() or f"{row.get('company_name')}|{row.get('event_headline')}".lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(row)

    return out


def add_coverage_depth(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    counts: dict[str, int] = {}

    for row in rows:
        company = row["company_name"]
        if is_last_12_months(row.get("event_date", "")):
            counts[company] = counts.get(company, 0) + 1

    for row in rows:
        row["coverage_depth_events_per_account_last_12mo"] = str(counts.get(row["company_name"], 0))

    return rows


def exa_rows(company: pd.Series, body: Any) -> list[dict[str, str]]:
    results = body.get("results", []) if isinstance(body, dict) else []
    rows = []

    for item in results:
        if not isinstance(item, dict):
            continue

        rows.append(
            normalize_event(
                company_name=clean(company["company_name"]),
                domain=clean(company["domain"]),
                company_linkedin_url=clean(company.get("linkedin_url")),
                title=item.get("title"),
                url=item.get("url"),
                date=item.get("publishedDate") or item.get("published_date"),
                publisher=item.get("author") or item.get("source") or "",
                summary=item.get("text") or item.get("summary") or "",
            )
        )

    return rows


def tavily_rows(company: pd.Series, body: Any) -> list[dict[str, str]]:
    results = body.get("results", []) if isinstance(body, dict) else []
    rows = []

    for item in results:
        if not isinstance(item, dict):
            continue

        rows.append(
            normalize_event(
                company_name=clean(company["company_name"]),
                domain=clean(company["domain"]),
                company_linkedin_url=clean(company.get("linkedin_url")),
                title=item.get("title"),
                url=item.get("url"),
                date=item.get("published_date") or item.get("date"),
                publisher=item.get("source") or "",
                summary=item.get("content") or item.get("raw_content") or "",
            )
        )

    return rows


def newsapi_rows(company: pd.Series, body: Any) -> list[dict[str, str]]:
    articles = body.get("articles", []) if isinstance(body, dict) else []
    rows = []

    for article in articles:
        if not isinstance(article, dict):
            continue

        source = article.get("source") if isinstance(article.get("source"), dict) else {}

        rows.append(
            normalize_event(
                company_name=clean(company["company_name"]),
                domain=clean(company["domain"]),
                company_linkedin_url=clean(company.get("linkedin_url")),
                title=article.get("title"),
                url=article.get("url"),
                date=article.get("publishedAt"),
                publisher=source.get("name") or article.get("author") or "",
                summary=article.get("description") or article.get("content") or "",
            )
        )

    return rows


def google_rss_rows(company: pd.Series, xml_text: str) -> list[dict[str, str]]:
    rows = []

    if not xml_text:
        return rows

    try:
        root = ET.fromstring(xml_text)
    except ET.ParseError:
        return rows

    channel = root.find("channel")
    if channel is None:
        return rows

    for item in channel.findall("item"):
        title = clean(item.findtext("title"))
        link = clean(item.findtext("link"))
        pub_date = clean(item.findtext("pubDate"))
        source = ""

        source_node = item.find("{https://news.google.com/rss}source")
        if source_node is not None:
            source = clean(source_node.text)

        rows.append(
            normalize_event(
                company_name=clean(company["company_name"]),
                domain=clean(company["domain"]),
                company_linkedin_url=clean(company.get("linkedin_url")),
                title=title,
                url=link,
                date=pub_date,
                publisher=source or "Google News RSS",
                summary=title,
            )
        )

    return rows


def missing_fields(rows: list[dict[str, str]]) -> list[str]:
    gated = []
    for field in DATA_COLUMNS:
        if field == "company_name":
            continue
        if any(not clean(row.get(field)) for row in rows):
            gated.append(field)
    return gated


def build_summary(
    *,
    tool_name: str,
    api_available: str,
    auth: str,
    credits: str,
    info: str,
    companies: pd.DataFrame,
    traces: list[dict[str, Any]],
    rows: list[dict[str, str]],
) -> dict[str, str]:

    statuses: dict[str, int] = {}
    for trace in traces:
        key = str(trace.get("status_code", ""))
        statuses[key] = statuses.get(key, 0) + 1

    success_count = sum(1 for trace in traces if trace.get("success"))
    error_count = len(traces) - success_count
    error_rate = f"{(error_count / len(traces) * 100):.2f}" if traces else "100.00"
    avg_latency = sum(float(trace.get("latency_sec") or 0) for trace in traces) / len(traces) if traces else 0
    records = len(rows)
    rate_limits = [clean(trace.get("rate_limit")) for trace in traces if clean(trace.get("rate_limit"))]
    gated = missing_fields(rows)

    return {
        "Tool Name": tool_name,
        "Category": "News / Key Announcements",
        "API Available (Y/N)": api_available,
        "Authentication Type": auth,
        "Credits Used": credits,
        "Rate Limit": rate_limits[0] if rate_limits else "Not returned in response headers",
        "Companies Processed": str(len(companies)),
        "Error Rate (%)": error_rate,
        "Latency (sec)": f"{avg_latency:.3f} avg",
        "Status Code / Error Flag": "; ".join(f"{k}:{v}" for k, v in sorted(statuses.items())) or "NO_CALLS",
        "Records Retrieved": str(records),
        "Gated Fields Flag": "Y - " + "; ".join(gated) if gated else "N",
        "missing_info": info,
        "Raw Export Saved (Y/N)": "Y",
    }


def write_tool_outputs(
    *,
    tool: str,
    tool_name: str,
    companies: pd.DataFrame,
    rows: list[dict[str, str]],
    traces: list[dict[str, Any]],
    api_available: str,
    auth: str,
    credits: str,
    info: str,
) -> None:

    tool_dir = OUT_ROOT / tool

    rows = dedupe_rows(rows)
    rows = add_coverage_depth(rows)

    summary = build_summary(
        tool_name=tool_name,
        api_available=api_available,
        auth=auth,
        credits=credits,
        info=info,
        companies=companies,
        traces=traces,
        rows=rows,
    )

    write_json(tool_dir / f"{tool}_data.json", rows)
    write_json(tool_dir / f"{tool}_api_records.json", [summary])
    write_json(tool_dir / f"{tool}_api_call_detail.json", traces)

    write_xlsx(tool_dir / f"{tool}_data.xlsx", pd.DataFrame(rows, columns=DATA_COLUMNS), f"{tool}_data")
    write_xlsx(tool_dir / f"{tool}_api_report.xlsx", pd.DataFrame([summary], columns=API_REPORT_COLUMNS), f"{tool}_api_report")


def run_exa(companies: pd.DataFrame, env: dict[str, str]) -> None:
    tool = "exa"
    tool_name = "Exa Search API"
    raw_dir = OUT_ROOT / tool / "raw_exports"
    api_key = env.get("EXA_API_KEY", "")

    rows: list[dict[str, str]] = []
    traces: list[dict[str, Any]] = []

    for _, company in companies.iterrows():
        company_name = clean(company["company_name"])
        domain = clean(company["domain"])
        raw_path = raw_dir / f"{safe_slug(company_name)}.json"

        if not api_key:
            body = {"error": "Missing EXA_API_KEY"}
            write_json(raw_path, body)
            trace = {
                "company_name": company_name,
                "status_code": "MISSING_KEY",
                "latency_sec": 0,
                "success": False,
                "records_retrieved": 0,
                "error": body["error"],
                "rate_limit": "",
            }
        else:
            headers = {
                "x-api-key": api_key,
                "Content-Type": "application/json",
            }

            payload = {
                "query": build_query(company_name, domain),
                "numResults": MAX_RESULTS_PER_COMPANY_PER_TOOL,
                "type": "auto",
                "contents": {
                    "text": True,
                    "highlights": True,
                },
                "startPublishedDate": (datetime.now(timezone.utc) - timedelta(days=LOOKBACK_DAYS)).date().isoformat(),
            }

            body, trace = request_json_api(
                "POST",
                "https://api.exa.ai/search",
                headers=headers,
                payload=payload,
                raw_path=raw_path,
            )

        traces.append({**trace, "company_name": company_name})
        rows.extend(exa_rows(company, body))

        time.sleep(0.5)

    write_tool_outputs(
        tool=tool,
        tool_name=tool_name,
        companies=companies,
        rows=rows,
        traces=traces,
        api_available="Y" if api_key else "N",
        auth="API Key in x-api-key header",
        credits="Search requests consume Exa account credits based on plan",
        info="Exa is used for web/news discovery with last-12-month filtering where provider dates are available.",
    )


def run_tavily(companies: pd.DataFrame, env: dict[str, str]) -> None:
    tool = "tavily"
    tool_name = "Tavily Search API"
    raw_dir = OUT_ROOT / tool / "raw_exports"
    api_key = env.get("TAVILY_API_KEY", "")

    rows: list[dict[str, str]] = []
    traces: list[dict[str, Any]] = []

    for _, company in companies.iterrows():
        company_name = clean(company["company_name"])
        domain = clean(company["domain"])
        raw_path = raw_dir / f"{safe_slug(company_name)}.json"

        if not api_key:
            body = {"error": "Missing TAVILY_API_KEY"}
            write_json(raw_path, body)
            trace = {
                "company_name": company_name,
                "status_code": "MISSING_KEY",
                "latency_sec": 0,
                "success": False,
                "records_retrieved": 0,
                "error": body["error"],
                "rate_limit": "",
            }
        else:
            headers = {
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json",
            }

            payload = {
                "query": build_query(company_name, domain),
                "topic": "news",
                "search_depth": "basic",
                "max_results": MAX_RESULTS_PER_COMPANY_PER_TOOL,
                "include_answer": False,
                "include_raw_content": False,
                "time_range": "year",
            }

            body, trace = request_json_api(
                "POST",
                "https://api.tavily.com/search",
                headers=headers,
                payload=payload,
                raw_path=raw_path,
            )

        traces.append({**trace, "company_name": company_name})
        rows.extend(tavily_rows(company, body))

        time.sleep(0.5)

    write_tool_outputs(
        tool=tool,
        tool_name=tool_name,
        companies=companies,
        rows=rows,
        traces=traces,
        api_available="Y" if api_key else "N",
        auth="Bearer token in Authorization header",
        credits="Basic search generally uses fewer credits than advanced search; exact usage depends on Tavily account plan",
        info="Tavily is used with topic=news and time_range=year for last-12-month announcement discovery.",
    )


def run_newsapi(companies: pd.DataFrame, env: dict[str, str]) -> None:
    tool = "newsapi"
    tool_name = "NewsAPI.org Everything API"
    raw_dir = OUT_ROOT / tool / "raw_exports"
    api_key = env.get("NEWSAPI_KEY", "")

    rows: list[dict[str, str]] = []
    traces: list[dict[str, Any]] = []

    from_date = (datetime.now(timezone.utc) - timedelta(days=LOOKBACK_DAYS)).date().isoformat()

    for _, company in companies.iterrows():
        company_name = clean(company["company_name"])
        domain = clean(company["domain"])
        raw_path = raw_dir / f"{safe_slug(company_name)}.json"

        if not api_key:
            body = {"error": "Missing NEWSAPI_KEY"}
            write_json(raw_path, body)
            trace = {
                "company_name": company_name,
                "status_code": "MISSING_KEY",
                "latency_sec": 0,
                "success": False,
                "records_retrieved": 0,
                "error": body["error"],
                "rate_limit": "",
            }
        else:
            headers = {
                "X-Api-Key": api_key,
            }

            params = {
                "q": build_query(company_name, domain),
                "from": from_date,
                "language": "en",
                "sortBy": "publishedAt",
                "pageSize": MAX_RESULTS_PER_COMPANY_PER_TOOL,
            }

            body, trace = request_json_api(
                "GET",
                "https://newsapi.org/v2/everything",
                headers=headers,
                params=params,
                raw_path=raw_path,
            )

        traces.append({**trace, "company_name": company_name})
        rows.extend(newsapi_rows(company, body))

        time.sleep(0.5)

    write_tool_outputs(
        tool=tool,
        tool_name=tool_name,
        companies=companies,
        rows=rows,
        traces=traces,
        api_available="Y" if api_key else "N",
        auth="API Key in X-Api-Key header",
        credits="NewsAPI request usage depends on account plan",
        info="NewsAPI Everything endpoint is used for article discovery over the last 12 months.",
    )


def run_google_news_rss(companies: pd.DataFrame) -> None:
    tool = "google_news_rss"
    tool_name = "Google News RSS"
    raw_dir = OUT_ROOT / tool / "raw_exports"

    rows: list[dict[str, str]] = []
    traces: list[dict[str, Any]] = []

    for _, company in companies.iterrows():
        company_name = clean(company["company_name"])
        domain = clean(company["domain"])

        query = build_query(company_name, domain)
        rss_url = f"https://news.google.com/rss/search?q={quote_plus(query)}&hl=en-US&gl=US&ceid=US:en"

        raw_path = raw_dir / f"{safe_slug(company_name)}.json"

        text, trace = request_text_api(
            "GET",
            rss_url,
            raw_path=raw_path,
            headers={"User-Agent": "Mozilla/5.0"},
        )

        traces.append({**trace, "company_name": company_name})
        rows.extend(google_rss_rows(company, text)[:MAX_RESULTS_PER_COMPANY_PER_TOOL])

        time.sleep(0.5)

    write_tool_outputs(
        tool=tool,
        tool_name=tool_name,
        companies=companies,
        rows=rows,
        traces=traces,
        api_available="Y",
        auth="No API key required",
        credits="0 API credits; public RSS fetch only",
        info="Google News RSS is unofficial/no-key RSS discovery. URLs may point to Google News redirect links instead of clean publisher URLs.",
    )


def main() -> None:
    companies = read_companies()

    if OUT_ROOT.exists():
        shutil.rmtree(OUT_ROOT)

    OUT_ROOT.mkdir(parents=True, exist_ok=True)

    env = load_env(ROOT / ".env")

    write_json(OUT_ROOT / "_input_snapshot.json", companies.to_dict(orient="records"))

    run_exa(companies, env)
    run_tavily(companies, env)
    run_newsapi(companies, env)
    run_google_news_rss(companies)

    print(f"Wrote {OUT_ROOT}")

    for path in sorted(OUT_ROOT.glob("*/*")):
        if path.is_file():
            print(path.relative_to(ROOT))


if __name__ == "__main__":
    main()
