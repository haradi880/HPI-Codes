from __future__ import annotations

import json
import re
import time
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent

INPUT = ROOT / "input" / "compnys.txt"
OUT_ROOT = ROOT / "ContactLevel"

HTTP_TIMEOUT = 60
CONTACTS_PER_COMPANY = 2


DATA_COLUMNS = [
    "company_name",
    "website_domain",
    "company_linkedin_url",
    "requested_contact",
    "matched_contact",
    "match_status",
    "match_rate_per_requested_contact",
    "verified_work_email",
    "email_confidence",
    "direct_mobile_phone",
    "phone_confidence",
    "title",
    "seniority",
    "department",
    "reports_to_org_chart",
    "linkedin_url",
]


API_REPORT_COLUMNS = [
    "Tool Name",
    "Category",
    "API Available (Y/N)",
    "Authentication Type",
    "Credits Used",
    "Rate Limit",
    "Companies Processed",
    "Contacts Requested",
    "Contacts Matched",
    "Match Rate (%)",
    "Error Rate (%)",
    "Latency (sec)",
    "Status Code / Error Flag",
    "Records Retrieved",
    "Gated Fields Flag",
    "missing_info",
    "Raw Export Saved (Y/N)",
]


SENIORITY_PATTERNS = [
    ("c_suite", re.compile(r"\b(chief|ceo|cfo|cio|cto|coo|cmo|ciso|president)\b", re.I)),
    ("vp", re.compile(r"\b(vp|vice president)\b", re.I)),
    ("director", re.compile(r"\b(director)\b", re.I)),
    ("head", re.compile(r"\b(head)\b", re.I)),
    ("manager", re.compile(r"\b(manager|lead)\b", re.I)),
    ("senior", re.compile(r"\b(senior|principal)\b", re.I)),
]


DEPARTMENT_PATTERNS = [
    ("technology", re.compile(r"\b(technology|engineering|software|data|digital|cloud|security|it|infrastructure|architecture|ai)\b", re.I)),
    ("finance", re.compile(r"\b(finance|financial|accounting|treasury|tax|risk|audit)\b", re.I)),
    ("sales", re.compile(r"\b(sales|business development|commercial|relationship|customer)\b", re.I)),
    ("operations", re.compile(r"\b(operations|supply|procurement|manufacturing|production|logistics)\b", re.I)),
    ("hr", re.compile(r"\b(hr|human resources|people|talent|recruit)\b", re.I)),
    ("legal", re.compile(r"\b(legal|compliance|counsel)\b", re.I)),
    ("marketing", re.compile(r"\b(marketing|brand|communications|growth)\b", re.I)),
]


@dataclass
class Seed:
    company_name: str
    domain: str
    company_linkedin_url: str
    full_name: str
    first_name: str
    last_name: str
    title: str
    seniority: str
    department: str
    linkedin_url: str
    apollo_id: str


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")


def safe_slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value).lower()).strip("_") or "item"


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    if isinstance(value, list):
        return "; ".join(clean(v) for v in value if clean(v))
    if isinstance(value, dict):
        return clean(first_value(value, ["name", "full_name", "value", "url", "number", "email"])) or json.dumps(value, ensure_ascii=False)[:200]
    text = str(value).strip()
    if text.lower() in {"none", "nan", "null", "not available", "n/a"}:
        return ""
    return text


def first_value(obj: dict[str, Any], keys: list[str]) -> Any:
    for key in keys:
        value = obj.get(key)
        if value not in (None, "", [], {}):
            return value
    return ""


def load_env() -> dict[str, str]:
    env: dict[str, str] = {}

    env_path = ROOT / ".env"
    if env_path.exists():
        for raw in env_path.read_text(encoding="utf-8").splitlines():
            line = raw.strip()
            if not line or line.startswith("#"):
                continue
            if "=" in line:
                key, value = raw.split("=", 1)
                env[key.strip()] = value.strip().strip('"').strip("'")
            elif line.startswith("202."):
                env["SIGNALHIRE_API_KEY"] = line

    api_path = ROOT / "Api.txt"
    if api_path.exists():
        for raw in api_path.read_text(encoding="utf-8").splitlines():
            parts = [part.strip() for part in raw.split(",")]
            if len(parts) < 4:
                continue

            tool = parts[0].lower()
            api = parts[3].strip()

            if "apollo" in tool:
                env["APOLLO_API_KEY"] = env.get("APOLLO_API_KEY") or api
            elif "fullenrich" in tool:
                env["FULLENRICH_API_KEY"] = env.get("FULLENRICH_API_KEY") or api
            elif "prospeo" in tool:
                env["PROSPEO_API_KEY"] = env.get("PROSPEO_API_KEY") or api
            elif "signalhire" in tool:
                env["SIGNALHIRE_API_KEY"] = env.get("SIGNALHIRE_API_KEY") or api

    return env


def read_companies() -> pd.DataFrame:
    if not INPUT.exists():
        raise FileNotFoundError(f"Input file not found: {INPUT}")

    companies = pd.read_csv(INPUT, dtype=str).fillna("")
    companies.columns = [c.strip() for c in companies.columns]

    required = ["company_name", "domain", "linkedin_url"]
    missing = [c for c in required if c not in companies.columns]
    if missing:
        raise ValueError(f"Missing required columns in compnys.txt: {missing}")

    companies["company_name"] = companies["company_name"].map(clean)
    companies["domain"] = companies["domain"].map(lambda x: clean(x).replace("https://", "").replace("http://", "").strip("/"))
    companies["linkedin_url"] = companies["linkedin_url"].map(clean)

    companies = companies[
        (companies["company_name"] != "") &
        (companies["domain"] != "")
    ].copy()

    return companies


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")


def format_workbook(path: Path) -> None:
    wb = load_workbook(path)
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for column in ws.columns:
            letter = get_column_letter(column[0].column)
            max_len = max((len(str(cell.value)) for cell in column if cell.value is not None), default=10)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 65)

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(path)


def write_xlsx(path: Path, rows: list[dict[str, Any]], columns: list[str], sheet_name: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        pd.DataFrame(rows, columns=columns).to_excel(writer, sheet_name=sheet_name, index=False)
    format_workbook(path)


def write_empty_contact_outputs(seed_report: list[dict[str, Any]]) -> None:
    tool_dir = OUT_ROOT / "no_seed_contacts"
    rows: list[dict[str, str]] = []
    report = {
        "Tool Name": "Contact-level seed search",
        "Category": "Contact Level",
        "API Available (Y/N)": "Y",
        "Authentication Type": "FullEnrich Bearer token; Apollo x-api-key fallback",
        "Credits Used": "No contact enrich credits used; seed searches were blocked before contact enrichment.",
        "Rate Limit": "",
        "Companies Processed": str(len(seed_report)),
        "Contacts Requested": "0",
        "Contacts Matched": "0",
        "Match Rate (%)": "0.00",
        "Error Rate (%)": "100.00",
        "Latency (sec)": f"{sum(float(item.get('latency_sec') or 0) for item in seed_report):.3f} total seed latency",
        "Status Code / Error Flag": "; ".join(
            filter(
                None,
                [
                    f"FullEnrich:{item.get('status_code')}" for item in seed_report
                ]
                + [
                    f"Apollo fallback:{item.get('fallback_status_code')}"
                    for item in seed_report
                    if item.get("fallback_status_code")
                ],
            )
        ),
        "Records Retrieved": "0",
        "Gated Fields Flag": "Y - seed contact search blocked by provider credits/plan; no person records available to enrich.",
        "missing_info": "; ".join(
            clean(item.get("error") or item.get("fallback_error"))
            for item in seed_report
            if clean(item.get("error") or item.get("fallback_error"))
        ),
        "Raw Export Saved (Y/N)": "Y",
    }
    write_json(tool_dir / "contact_level_data.json", rows)
    write_json(tool_dir / "contact_level_api_records.json", [report])
    write_xlsx(tool_dir / "contact_level_data.xlsx", rows, DATA_COLUMNS, "contact_level_data")
    write_xlsx(tool_dir / "contact_level_api_report.xlsx", [report], API_REPORT_COLUMNS, "contact_level_api_report")


def response_json(response: requests.Response) -> Any:
    try:
        return response.json()
    except ValueError:
        return {"raw_text": response.text}


def save_response(path: Path, request_data: dict[str, Any], response: requests.Response, latency_sec: float) -> Any:
    body = response_json(response)
    write_json(
        path,
        {
            "request": request_data,
            "response": {
                "status_code": response.status_code,
                "headers": dict(response.headers),
                "latency_sec": round(latency_sec, 3),
                "text": response.text if not isinstance(body, (dict, list)) else "",
            },
            "json": body,
        },
    )
    return body


def request_api(
    method: str,
    url: str,
    *,
    headers: dict[str, str],
    raw_path: Path,
    params: dict[str, Any] | None = None,
    payload: dict[str, Any] | None = None,
) -> tuple[Any, dict[str, Any]]:

    started = time.perf_counter()

    try:
        response = requests.request(
            method,
            url,
            headers=headers,
            params=params,
            json=payload,
            timeout=HTTP_TIMEOUT,
        )
        latency = time.perf_counter() - started
        body = save_response(
            raw_path,
            {"method": method, "url": url, "params": params or {}, "payload": payload or {}},
            response,
            latency,
        )

        trace = {
            "status_code": response.status_code,
            "latency_sec": round(latency, 3),
            "success": response.ok,
            "error": "" if response.ok else response.text[:500],
            "rate_limit": "; ".join(
                f"{k}: {v}"
                for k, v in response.headers.items()
                if "rate" in k.lower() or "credit" in k.lower()
            ),
        }
        return body, trace

    except Exception as exc:
        latency = time.perf_counter() - started
        body = {"error": str(exc)}
        write_json(
            raw_path,
            {
                "request": {"method": method, "url": url, "params": params or {}, "payload": payload or {}},
                "response": {"status_code": "EXCEPTION", "latency_sec": round(latency, 3)},
                "json": body,
            },
        )
        return body, {
            "status_code": "EXCEPTION",
            "latency_sec": round(latency, 3),
            "success": False,
            "error": str(exc),
            "rate_limit": "",
        }


def flatten_dicts(payload: Any) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []

    if isinstance(payload, dict):
        out.append(payload)
        for value in payload.values():
            if isinstance(value, (dict, list)):
                out.extend(flatten_dicts(value))

    elif isinstance(payload, list):
        for item in payload:
            out.extend(flatten_dicts(item))

    return out


def find_people(payload: Any) -> list[dict[str, Any]]:
    if isinstance(payload, list):
        return [item for item in payload if isinstance(item, dict)]

    if not isinstance(payload, dict):
        return []

    for key in ["people", "persons", "contacts", "results", "data", "items", "profiles"]:
        value = payload.get(key)
        if isinstance(value, list):
            return [item for item in value if isinstance(item, dict)]

    for value in payload.values():
        if isinstance(value, (dict, list)):
            found = find_people(value)
            if found:
                return found

    return []


def infer_seniority(title: str) -> str:
    for label, pattern in SENIORITY_PATTERNS:
        if pattern.search(title):
            return label
    return ""


def infer_department(title: str) -> str:
    for label, pattern in DEPARTMENT_PATTERNS:
        if pattern.search(title):
            return label
    return ""


def find_person_linkedin(payload: Any) -> str:
    for candidate in flatten_dicts(payload):
        for key in ["linkedin_url", "linkedin", "linkedin_profile_url", "url", "link"]:
            value = clean(candidate.get(key))
            if "linkedin.com/in/" in value.lower():
                return value
    return ""


def normalize_seed(company_name: str, domain: str, company_linkedin_url: str, person: dict[str, Any]) -> Seed:
    full_name = clean(first_value(person, ["name", "full_name", "display_name", "fullName"]))
    first_name = clean(first_value(person, ["first_name", "firstname", "firstName"]))
    last_name = clean(first_value(person, ["last_name", "lastname", "lastName"]))

    if not full_name:
        full_name = " ".join(part for part in [first_name, last_name] if part)

    if full_name and not first_name:
        pieces = full_name.split()
        first_name = pieces[0]
        last_name = " ".join(pieces[1:])

    employment = person.get("employment") if isinstance(person.get("employment"), dict) else {}
    current = employment.get("current") if isinstance(employment.get("current"), dict) else {}

    title = (
        clean(first_value(person, ["title", "job_title", "headline", "position"]))
        or clean(first_value(current, ["title"]))
    )

    return Seed(
        company_name=company_name,
        domain=domain,
        company_linkedin_url=company_linkedin_url,
        full_name=full_name,
        first_name=first_name,
        last_name=last_name,
        title=title,
        seniority=clean(first_value(person, ["seniority", "seniority_level"])) or infer_seniority(title),
        department=clean(first_value(person, ["department", "departments", "function"])) or infer_department(title),
        linkedin_url=find_person_linkedin(person),
        apollo_id=clean(first_value(person, ["id", "person_id"])),
    )


def seed_key(seed: Seed) -> str:
    return (seed.linkedin_url or f"{seed.company_name}|{seed.full_name}|{seed.title}").lower()


def build_seeds(companies: pd.DataFrame, env: dict[str, str]) -> list[Seed]:
    seed_dir = OUT_ROOT / "_seed_contacts"
    raw_dir = seed_dir / "raw_exports"

    seeds: list[Seed] = []
    seed_report: list[dict[str, Any]] = []

    api_key = env.get("FULLENRICH_API_KEY", "")
    apollo_api_key = env.get("APOLLO_API_KEY", "")

    headers = {
        "accept": "application/json",
        "Content-Type": "application/json",
        "Authorization": f"Bearer {api_key}",
    }

    for _, company in companies.iterrows():
        company_name = clean(company["company_name"])
        domain = clean(company["domain"])
        company_linkedin_url = clean(company.get("linkedin_url", ""))

        raw_path = raw_dir / f"{safe_slug(company_name)}.json"

        if not api_key:
            body = {"error": "Missing FULLENRICH_API_KEY. Seed contacts cannot be fetched."}
            write_json(raw_path, body)

            trace = {
                "status_code": "MISSING_KEY",
                "latency_sec": 0,
                "success": False,
                "error": body["error"],
                "rate_limit": "",
            }
            people = []

        else:
            payload = {
                "offset": 0,
                "limit": CONTACTS_PER_COMPANY,
                "current_company_domains": [
                    {
                        "value": domain,
                        "exact_match": True,
                        "exclude": False,
                    }
                ],
                "current_position_seniority_level": [
                    {"value": "C-level", "exact_match": False, "exclude": False},
                    {"value": "VP", "exact_match": False, "exclude": False},
                    {"value": "Director", "exact_match": False, "exclude": False},
                    {"value": "Manager", "exact_match": False, "exclude": False},
                ],
            }

            body, trace = request_api(
                "POST",
                "https://app.fullenrich.com/api/v2/people/search",
                headers=headers,
                payload=payload,
                raw_path=raw_path,
            )

            people = find_people(body)[:CONTACTS_PER_COMPANY]

            if not people and apollo_api_key:
                apollo_headers = {
                    "accept": "application/json",
                    "Cache-Control": "no-cache",
                    "Content-Type": "application/json",
                    "x-api-key": apollo_api_key,
                }
                apollo_payload = {
                    "page": 1,
                    "per_page": CONTACTS_PER_COMPANY,
                    "q_organization_domains": domain,
                    "person_seniorities": ["c_suite", "vp", "director", "manager"],
                }
                apollo_body, apollo_trace = request_api(
                    "POST",
                    "https://api.apollo.io/api/v1/mixed_people/search",
                    headers=apollo_headers,
                    payload=apollo_payload,
                    raw_path=raw_dir / f"{safe_slug(company_name)}__apollo_seed_fallback.json",
                )
                apollo_people = find_people(apollo_body)[:CONTACTS_PER_COMPANY]
                if apollo_people:
                    people = apollo_people
                    trace = {
                        **apollo_trace,
                        "fallback_from": f"FullEnrich seed search returned {trace.get('status_code')}",
                        "seed_source": "apollo",
                    }
                else:
                    trace = {
                        **trace,
                        "fallback_status_code": apollo_trace.get("status_code"),
                        "fallback_error": apollo_trace.get("error", ""),
                        "seed_source": "fullenrich",
                    }

        seen: set[str] = set()

        for person in people:
            seed = normalize_seed(company_name, domain, company_linkedin_url, person)

            if seed.full_name and seed_key(seed) not in seen:
                seeds.append(seed)
                seen.add(seed_key(seed))

        seed_report.append(
            {
                **trace,
                "company_name": company_name,
                "domain": domain,
                "company_linkedin_url": company_linkedin_url,
                "records_retrieved": len(people),
            }
        )

    seed_rows = [seed.__dict__ for seed in seeds]

    write_json(seed_dir / "seed_contacts.json", seed_rows)
    write_json(seed_dir / "seed_api_call_detail.json", seed_report)

    if seed_rows:
        write_xlsx(seed_dir / "seed_contacts.xlsx", seed_rows, list(seed_rows[0].keys()), "seed_contacts")

    return seeds


def normalize_email(value: Any) -> tuple[str, str]:
    if isinstance(value, str):
        return value if "@" in value else "", "provided" if "@" in value else ""

    if isinstance(value, dict):
        email = clean(first_value(value, ["email", "value", "address", "email_address"]))
        confidence = clean(first_value(value, ["confidence", "status", "verification_status", "type", "quality"]))
        return (email if "@" in email else ""), confidence

    if isinstance(value, list):
        for item in value:
            email, confidence = normalize_email(item)
            if email:
                return email, confidence

    return "", ""


def normalize_phone(value: Any) -> tuple[str, str]:
    if isinstance(value, str):
        return value, "provided" if value else ""

    if isinstance(value, dict):
        phone = clean(first_value(value, ["phone", "number", "value", "phone_number", "sanitized_number"]))
        confidence = clean(first_value(value, ["confidence", "status", "type", "quality"]))
        return phone, confidence

    if isinstance(value, list):
        for item in value:
            phone, confidence = normalize_phone(item)
            if phone:
                return phone, confidence

    return "", ""


def find_first_email(obj: dict[str, Any]) -> tuple[str, str]:
    for candidate in flatten_dicts(obj):
        for key in [
            "verified_work_email",
            "most_probable_email",
            "work_email",
            "verified_email",
            "professional_email",
            "email",
            "email_address",
            "emails",
            "work_emails",
        ]:
            if key in candidate:
                email, confidence = normalize_email(candidate.get(key))
                if email:
                    return email, confidence or "returned"
    return "", ""


def find_first_phone(obj: dict[str, Any]) -> tuple[str, str]:
    for candidate in flatten_dicts(obj):
        for key in [
            "direct_mobile_phone",
            "most_probable_phone",
            "mobile_phone",
            "phone",
            "phone_number",
            "direct_phone",
            "phones",
            "mobile_phones",
        ]:
            if key in candidate:
                phone, confidence = normalize_phone(candidate.get(key))
                if phone:
                    return phone, confidence or "returned"
    return "", ""


def choose_best(seed: Seed, payload: Any) -> dict[str, Any]:
    candidates = flatten_dicts(payload)

    clean_candidates = []
    for candidate in candidates:
        if candidate.get("error") or candidate.get("error_code"):
            continue
        clean_candidates.append(candidate)

    target = seed.full_name.lower().strip()

    for candidate in clean_candidates:
        linkedin = clean(first_value(candidate, ["linkedin_url", "linkedin", "linkedin_profile_url", "url"])).lower().rstrip("/")
        if seed.linkedin_url and seed.linkedin_url.lower().rstrip("/") in linkedin:
            return candidate

    for candidate in clean_candidates:
        name = clean(first_value(candidate, ["name", "full_name", "fullName", "display_name"])).lower()
        if target and name and (target in name or name in target):
            return candidate

    for candidate in clean_candidates:
        email, _ = find_first_email(candidate)
        phone, _ = find_first_phone(candidate)
        if email or phone:
            return candidate

    return clean_candidates[0] if clean_candidates else {}


def row_from_payload(seed: Seed, payload: Any) -> dict[str, str]:
    best = choose_best(seed, payload)

    email, email_conf = find_first_email(best)
    phone, phone_conf = find_first_phone(best)

    title = (
        clean(first_value(best, ["title", "job_title", "headline", "headLine", "position"]))
        or seed.title
    )

    matched = clean(first_value(best, ["name", "full_name", "fullName", "display_name"]))

    if not matched and best:
        matched = " ".join(
            part
            for part in [
                clean(first_value(best, ["firstname", "first_name"])),
                clean(first_value(best, ["lastname", "last_name"])),
            ]
            if part
        )

    match_status = "Matched" if best else "Not matched"

    return {
        "company_name": seed.company_name,
        "website_domain": seed.domain,
        "company_linkedin_url": seed.company_linkedin_url,
        "requested_contact": seed.full_name,
        "matched_contact": matched,
        "match_status": match_status,
        "match_rate_per_requested_contact": "100.00" if match_status == "Matched" else "0.00",
        "verified_work_email": email,
        "email_confidence": email_conf,
        "direct_mobile_phone": phone,
        "phone_confidence": phone_conf,
        "title": title,
        "seniority": clean(first_value(best, ["seniority", "seniority_level", "management_level"])) or seed.seniority or infer_seniority(title),
        "department": clean(first_value(best, ["department", "departments", "function"])) or seed.department or infer_department(title),
        "reports_to_org_chart": clean(first_value(best, ["reports_to", "manager", "manager_name", "org_chart", "hierarchy"])),
        "linkedin_url": find_person_linkedin(best) or seed.linkedin_url,
    }


def traces_summary(
    tool_name: str,
    traces: list[dict[str, Any]],
    rows: list[dict[str, str]],
    auth: str,
    credits: str,
    info: str,
) -> dict[str, str]:

    statuses: dict[str, int] = {}
    for trace in traces:
        code = str(trace.get("status_code", ""))
        statuses[code] = statuses.get(code, 0) + 1

    requested = len(rows)
    matched = sum(1 for row in rows if row["match_status"] == "Matched")
    successes = sum(1 for trace in traces if trace.get("success"))

    missing_fields = []
    for field in ["verified_work_email", "direct_mobile_phone", "title", "reports_to_org_chart", "linkedin_url"]:
        if any(not clean(row.get(field)) for row in rows):
            missing_fields.append(field)

    rates = [clean(trace.get("rate_limit")) for trace in traces if clean(trace.get("rate_limit"))]
    latency = sum(float(trace.get("latency_sec") or 0) for trace in traces) / len(traces) if traces else 0

    return {
        "Tool Name": tool_name,
        "Category": "Contact-level",
        "API Available (Y/N)": "Y",
        "Authentication Type": auth,
        "Credits Used": credits,
        "Rate Limit": rates[0] if rates else "Not returned in response headers",
        "Companies Processed": str(len({row["company_name"] for row in rows})),
        "Contacts Requested": str(requested),
        "Contacts Matched": str(matched),
        "Match Rate (%)": f"{(matched / requested * 100) if requested else 0:.2f}",
        "Error Rate (%)": f"{((len(traces) - successes) / len(traces) * 100) if traces else 100:.2f}",
        "Latency (sec)": f"{latency:.3f} avg",
        "Status Code / Error Flag": "; ".join(f"{k}:{v}" for k, v in sorted(statuses.items())),
        "Records Retrieved": str(matched),
        "Gated Fields Flag": "Y - " + "; ".join(missing_fields) if missing_fields else "N",
        "missing_info": info,
        "Raw Export Saved (Y/N)": "Y",
    }


def write_tool_outputs(
    tool: str,
    tool_name: str,
    rows: list[dict[str, str]],
    traces: list[dict[str, Any]],
    auth: str,
    credits: str,
    info: str,
) -> None:

    tool_dir = OUT_ROOT / tool
    report = traces_summary(tool_name, traces, rows, auth, credits, info)

    write_json(tool_dir / f"{tool}_data.json", rows)
    write_json(tool_dir / f"{tool}_api_records.json", [report])
    write_json(tool_dir / f"{tool}_api_call_detail.json", traces)

    write_xlsx(tool_dir / f"{tool}_data.xlsx", rows, DATA_COLUMNS, f"{tool}_data")
    write_xlsx(tool_dir / f"{tool}_api_report.xlsx", [report], API_REPORT_COLUMNS, f"{tool}_api_report")


def run_apollo(seeds: list[Seed], env: dict[str, str]) -> None:
    tool = "apollo"
    raw_dir = OUT_ROOT / tool / "raw_exports"
    api_key = env.get("APOLLO_API_KEY", "")

    rows = []
    traces = []

    for seed in seeds:
        raw_path = raw_dir / f"{safe_slug(seed.company_name)}__{safe_slug(seed.full_name)}.json"

        if not api_key:
            body = {"error": "Missing APOLLO_API_KEY"}
            write_json(raw_path, body)
            trace = {"status_code": "MISSING_KEY", "latency_sec": 0, "success": False, "error": body["error"], "rate_limit": ""}
        else:
            params = {
                "name": seed.full_name,
                "domain": seed.domain,
                "reveal_personal_emails": "false",
                "reveal_phone_number": "true",
            }

            if seed.linkedin_url:
                params["linkedin_url"] = seed.linkedin_url

            headers = {
                "accept": "application/json",
                "Content-Type": "application/json",
                "Cache-Control": "no-cache",
                "x-api-key": api_key,
            }

            body, trace = request_api(
                "POST",
                "https://api.apollo.io/api/v1/people/match",
                headers=headers,
                params=params,
                raw_path=raw_path,
            )

        traces.append({**trace, "company_name": seed.company_name})
        rows.append(row_from_payload(seed, body))

        time.sleep(0.3)

    write_tool_outputs(
        tool,
        "Apollo People Enrichment API",
        rows,
        traces,
        "API Key in x-api-key header",
        "Apollo people enrichment / phone reveal credits depend on plan",
        "Apollo matched contacts from seed names, domains and LinkedIn URLs. Missing email/phone means data unavailable or gated by plan.",
    )


def run_prospeo(seeds: list[Seed], env: dict[str, str]) -> None:
    tool = "prospeo"
    raw_dir = OUT_ROOT / tool / "raw_exports"
    api_key = env.get("PROSPEO_API_KEY", "")

    rows = []
    traces = []

    for seed in seeds:
        raw_path = raw_dir / f"{safe_slug(seed.company_name)}__{safe_slug(seed.full_name)}.json"

        if not api_key:
            body = {"error": "Missing PROSPEO_API_KEY"}
            write_json(raw_path, body)
            trace = {"status_code": "MISSING_KEY", "latency_sec": 0, "success": False, "error": body["error"], "rate_limit": ""}
        else:
            headers = {
                "accept": "application/json",
                "Content-Type": "application/json",
                "X-KEY": api_key,
            }

            payload = {
                "only_verified_email": True,
                "enrich_mobile": True,
                "data": {
                    "first_name": seed.first_name,
                    "last_name": seed.last_name,
                    "full_name": seed.full_name,
                    "company": seed.company_name,
                    "company_name": seed.company_name,
                    "company_website": seed.domain,
                    "linkedin_url": seed.linkedin_url,
                },
            }

            body, trace = request_api(
                "POST",
                "https://api.prospeo.io/enrich-person",
                headers=headers,
                payload=payload,
                raw_path=raw_path,
            )

        traces.append({**trace, "company_name": seed.company_name})
        rows.append(row_from_payload(seed, body))

        time.sleep(1.2)

    write_tool_outputs(
        tool,
        "Prospeo Person Enrichment API",
        rows,
        traces,
        "API Key in X-KEY header",
        "Prospeo enrichment credits depend on account plan",
        "Prospeo enriches by name, domain and LinkedIn URL. Missing values are provider gaps or gated fields.",
    )


def run_fullenrich(seeds: list[Seed], env: dict[str, str]) -> None:
    tool = "fullenrich"
    raw_dir = OUT_ROOT / tool / "raw_exports"
    api_key = env.get("FULLENRICH_API_KEY", "")

    rows = []
    traces = []

    for company_name in sorted({seed.company_name for seed in seeds}):
        group = [seed for seed in seeds if seed.company_name == company_name]
        raw_path = raw_dir / f"{safe_slug(company_name)}.json"

        if not api_key:
            body = {"error": "Missing FULLENRICH_API_KEY"}
            write_json(raw_path, body)
            trace = {"status_code": "MISSING_KEY", "latency_sec": 0, "success": False, "error": body["error"], "rate_limit": ""}
            final_body = body

        else:
            headers = {
                "accept": "application/json",
                "Content-Type": "application/json",
                "Authorization": f"Bearer {api_key}",
            }

            payload = {
                "name": f"HPI ContactLevel {company_name} {now_iso()}",
                "data": [
                    {
                        "first_name": seed.first_name,
                        "last_name": seed.last_name,
                        "domain": seed.domain,
                        "company_name": seed.company_name,
                        "linkedin_url": seed.linkedin_url,
                        "enrich_fields": ["contact.work_emails", "contact.phones"],
                        "custom": {"requested_contact": seed.full_name},
                    }
                    for seed in group
                ],
            }

            body, trace = request_api(
                "POST",
                "https://app.fullenrich.com/api/v2/contact/enrich/bulk",
                headers=headers,
                payload=payload,
                raw_path=raw_path,
            )

            final_body = body

        traces.append({**trace, "company_name": company_name})

        for seed in group:
            rows.append(row_from_payload(seed, final_body))

    write_tool_outputs(
        tool,
        "FullEnrich Contact Bulk Enrichment API",
        rows,
        traces,
        "Bearer token in Authorization header",
        "FullEnrich waterfall credits depend on account plan",
        "FullEnrich bulk enrichment used seed contacts generated from the company domain input.",
    )


def run_signalhire(seeds: list[Seed], env: dict[str, str]) -> None:
    tool = "signalhire"
    raw_dir = OUT_ROOT / tool / "raw_exports"
    api_key = env.get("SIGNALHIRE_API_KEY", "")

    rows = []
    traces = []

    selected = [seed for seed in seeds if seed.linkedin_url]

    for seed in selected:
        raw_path = raw_dir / f"{safe_slug(seed.company_name)}__{safe_slug(seed.full_name)}.json"

        if not api_key:
            body = {"error": "Missing SIGNALHIRE_API_KEY"}
            write_json(raw_path, body)
            trace = {"status_code": "MISSING_KEY", "latency_sec": 0, "success": False, "error": body["error"], "rate_limit": ""}
        else:
            headers = {
                "accept": "application/json",
                "Content-Type": "application/json",
                "apikey": api_key,
            }

            payload = {
                "items": [seed.linkedin_url],
                "withoutWaterfall": True,
            }

            body, trace = request_api(
                "POST",
                "https://www.signalhire.com/api/v1/candidate/search",
                headers=headers,
                payload=payload,
                raw_path=raw_path,
            )

        traces.append({**trace, "company_name": seed.company_name})
        rows.append(row_from_payload(seed, body))

        time.sleep(0.5)

    write_tool_outputs(
        tool,
        "SignalHire Candidate Search API",
        rows,
        traces,
        "API Key in apikey header",
        "SignalHire credits depend on account plan",
        "SignalHire only runs for seed contacts that have a personal LinkedIn profile URL.",
    )


def main() -> None:
    print("Reading companies from:", INPUT)

    env = load_env()
    companies = read_companies()

    OUT_ROOT.mkdir(parents=True, exist_ok=True)

    write_json(
        OUT_ROOT / "_input_snapshot.json",
        companies.to_dict(orient="records"),
    )

    print(f"Companies loaded: {len(companies)}")

    seeds = build_seeds(companies, env)

    print(f"Seed contacts found: {len(seeds)}")

    if not seeds:
        seed_report_path = OUT_ROOT / "_seed_contacts" / "seed_api_call_detail.json"
        seed_report = json.loads(seed_report_path.read_text(encoding="utf-8")) if seed_report_path.exists() else []
        write_empty_contact_outputs(seed_report)
        print("No seed contacts found. Check FULLENRICH_API_KEY or seed search response.")
        return

    run_apollo(seeds, env)
    run_fullenrich(seeds, env)
    run_prospeo(seeds, env)
    run_signalhire(seeds, env)

    print(f"Wrote contact-level outputs to: {OUT_ROOT}")


if __name__ == "__main__":
    main()
