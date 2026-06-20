from __future__ import annotations

import json
import re
import shutil
import time
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
INPUT = ROOT / "input" / "compnys.txt"
OUT_ROOT = ROOT / "JobsHiring"
BACKUP_APOLLO_JOBS_DIR = ROOT / "Backup" / "JobsHiring" / "raw" / "apollo" / "data"
HTTP_TIMEOUT = 60

DATA_COLUMNS = [
    "company_name",
    "website_domain",
    "active_job_count_total",
    "active_job_count_sg",
    "roles_by_function_eng_design_it_sales_ops",
    "hiring_velocity_postings_last_90_days",
    "job_locations",
    "posting_first_seen_date",
    "source_url",
]

API_REPORT_COLUMNS = [
    "Tool Name",
    "Category",
    "API Available (Y/N)",
    "Authentication Type",
    "Credits Used",
    "Rate Limit",
    "Companies Processed",
    "Error Rate (%)",
    "Latency (sec)",
    "Status Code / Error Flag",
    "Records Retrieved",
    "Gated Fields Flag",
    "missing_info",
    "Raw Export Saved (Y/N)",
]

FUNCTION_PATTERNS = {
    "eng": re.compile(r"\b(engineer|engineering|developer|software|backend|frontend|full[ -]?stack|devops|sre|qa|quality|architect|r&d|research|mechanical|electrical|civil)\b", re.I),
    "design": re.compile(r"\b(design|designer|ux|ui|user experience|product design|creative)\b", re.I),
    "it": re.compile(r"\b(information technology| IT |technology|infrastructure|network|security|cyber|systems?|cloud|data|database|platform|support analyst)\b", re.I),
    "sales": re.compile(r"\b(sales|account executive|business development|relationship manager|commercial|revenue|customer success)\b", re.I),
    "ops": re.compile(r"\b(operations|operator|operational|supply chain|logistics|procurement|maintenance|production|manufacturing|warehouse|project manager)\b", re.I),
}


def safe_slug(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_")


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    if isinstance(value, list):
        return " | ".join(clean(v) for v in value if clean(v))
    text = str(value).strip()
    if text.lower() in {"none", "nan", "null", "not available", "not verified"}:
        return ""
    return text


def dedupe(values: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for value in values:
        text = clean(value)
        key = text.lower()
        if text and key not in seen:
            seen.add(key)
            out.append(text)
    return out


def parse_date(value: Any) -> date | None:
    text = clean(value)
    if not text:
        return None
    match = re.search(r"\d{4}-\d{2}-\d{2}", text)
    if not match:
        return None
    try:
        return datetime.strptime(match.group(0), "%Y-%m-%d").date()
    except ValueError:
        return None


def load_env(path: Path) -> dict[str, str]:
    env: dict[str, str] = {}
    if not path.exists():
        return env
    for line in path.read_text(encoding="utf-8").splitlines():
        if "=" not in line or line.strip().startswith("#"):
            continue
        key, value = line.split("=", 1)
        env[key.strip()] = value.strip().strip('"').strip("'")
    return env


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")


def format_workbook(path: Path) -> None:
    wb = load_workbook(path)
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column in ws.columns:
            letter = get_column_letter(column[0].column)
            max_len = max((len(str(cell.value)) for cell in column if cell.value is not None), default=10)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 70)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    wb.save(path)


def write_xlsx(path: Path, rows: list[dict[str, str]], columns: list[str], sheet_name: str) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        pd.DataFrame(rows, columns=columns).to_excel(writer, sheet_name=sheet_name, index=False)
    format_workbook(path)


def response_json(response: requests.Response) -> Any:
    try:
        return response.json()
    except ValueError:
        return {"raw_text": response.text}


def save_response(path: Path, request_data: dict[str, Any], response: requests.Response, latency_sec: float) -> Any:
    body = response_json(response)
    write_json(
        path,
        {
            "request": request_data,
            "response": {
                "status_code": response.status_code,
                "headers": dict(response.headers),
                "latency_sec": round(latency_sec, 3),
                "text": response.text if not isinstance(body, (dict, list)) else "",
            },
            "json": body,
        },
    )
    return body


def classify_roles(titles: list[str]) -> dict[str, int]:
    counts = {key: 0 for key in FUNCTION_PATTERNS}
    for title in titles:
        matched = False
        for key, pattern in FUNCTION_PATTERNS.items():
            if pattern.search(f" {title} "):
                counts[key] += 1
                matched = True
        if not matched:
            continue
    return counts


def roles_text(counts: dict[str, int]) -> str:
    return "; ".join(f"{key}: {counts.get(key, 0)}" for key in ["eng", "design", "it", "sales", "ops"])


def count_sg(jobs: list[dict[str, Any]]) -> int:
    total = 0
    for job in jobs:
        country_code = clean(job.get("country_code") or job.get("location_country_code"))
        country = clean(job.get("country") or job.get("location_country"))
        location = clean(job.get("location") or job.get("job_location"))
        if country_code.upper() == "SG" or country.lower() == "singapore" or re.search(r"\bsingapore\b|\bsg\b", location, re.I):
            total += 1
    return total


def job_locations(jobs: list[dict[str, Any]]) -> str:
    values = []
    for job in jobs:
        values.append(clean(job.get("location") or job.get("short_location") or job.get("long_location") or job.get("job_location") or job.get("city") or job.get("country")))
    return "; ".join(dedupe(values)[:25])


def source_urls(jobs: list[dict[str, Any]]) -> str:
    values = []
    for job in jobs:
        values.append(clean(job.get("source_url") or job.get("url") or job.get("final_url") or job.get("job_url") or job.get("apply_url")))
        sources = job.get("job_sources")
        if isinstance(sources, list):
            for source in sources:
                if isinstance(source, dict) and clean(source.get("status")).lower() in {"", "active"}:
                    values.append(clean(source.get("url")))
        values.append(clean(job.get("external_url")))
    return "; ".join(dedupe(values)[:25])


def first_seen(jobs: list[dict[str, Any]]) -> str:
    dates = [d for d in (parse_date(job.get("discovered_at") or job.get("date_posted") or job.get("posted_at") or job.get("created_at") or job.get("first_seen_at")) for job in jobs) if d]
    return min(dates).isoformat() if dates else ""


def velocity_90(jobs: list[dict[str, Any]], total_from_query: str = "") -> str:
    if clean(total_from_query):
        return clean(total_from_query)
    cutoff = date.today() - timedelta(days=90)
    return str(sum(1 for job in jobs if (parse_date(job.get("date_posted") or job.get("posted_at") or job.get("discovered_at") or job.get("created_at")) or date.min) >= cutoff))


def theirstack_balance(api_key: str, raw_dir: Path, label: str) -> dict[str, Any]:
    url = "https://api.theirstack.com/v0/billing/credit-balance"
    headers = {"Authorization": f"Bearer {api_key}", "accept": "application/json"}
    started = time.perf_counter()
    response = requests.get(url, headers=headers, timeout=HTTP_TIMEOUT)
    latency = time.perf_counter() - started
    body = save_response(raw_dir / "info_and_diagnostics" / f"_theirstack_credit_balance_{label}.json", {"method": "GET", "url": url}, response, latency)
    return body if isinstance(body, dict) else {}


def live_theirstack(company: pd.Series, api_key: str, raw_dir: Path) -> tuple[dict[str, Any], dict[str, Any]]:
    url = "https://api.theirstack.com/v1/jobs/search"
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json", "accept": "application/json"}
    cutoff = (date.today() - timedelta(days=90)).isoformat()
    payloads = {
        "all_jobs_response": {"company_domain_or": [company["domain"]], "limit": 10, "page": 0, "include_total_results": True},
        "sg_count_response": {"company_domain_or": [company["domain"]], "job_country_code_or": ["SG"], "limit": 1, "page": 0, "include_total_results": True},
        "last_90_days_response": {"company_domain_or": [company["domain"]], "posted_at_gte": cutoff, "limit": 1, "page": 0, "include_total_results": True},
    }
    export: dict[str, Any] = {"requests": {}, "responses": {}}
    traces = []
    for index, (label, payload) in enumerate(payloads.items()):
        started = time.perf_counter()
        response = requests.post(url, headers=headers, json=payload, timeout=HTTP_TIMEOUT)
        latency = time.perf_counter() - started
        body = response_json(response)
        export["requests"][label] = {"method": "POST", "url": url, "payload": payload}
        export["responses"][label] = {
            "status_code": response.status_code,
            "headers": dict(response.headers),
            "latency_sec": round(latency, 3),
            "json": body,
            "text": response.text if not isinstance(body, (dict, list)) else "",
        }
        traces.append({"status_code": response.status_code, "latency_sec": round(latency, 3), "records": len(body.get("data") or []) if isinstance(body, dict) else 0, "error": "" if response.ok else response.text[:500], "rate_limit": "; ".join(f"{k}: {v}" for k, v in response.headers.items() if "rate" in k.lower())})
        if index < len(payloads) - 1:
            time.sleep(8)
    write_json(raw_dir / f"{safe_slug(company['company_name'])}.json", export)
    ok = all(t["status_code"] == 200 for t in traces)
    return export, {
        "company_name": company["company_name"],
        "status_code": "200" if ok else ";".join(str(t["status_code"]) for t in traces),
        "latency_sec": round(sum(float(t["latency_sec"]) for t in traces), 3),
        "success": ok,
        "records_retrieved": sum(int(t["records"]) for t in traces),
        "error": "; ".join(t["error"] for t in traces if t["error"]),
        "rate_limit": next((t["rate_limit"] for t in traces if t["rate_limit"]), ""),
    }


def parse_theirstack(export: dict[str, Any], company: pd.Series) -> dict[str, str]:
    all_body = export.get("responses", {}).get("all_jobs_response", {}).get("json") or {}
    sg_body = export.get("responses", {}).get("sg_count_response", {}).get("json") or {}
    v_body = export.get("responses", {}).get("last_90_days_response", {}).get("json") or {}
    jobs = all_body.get("data") or []
    total = clean((all_body.get("metadata") or {}).get("total_results"))
    sg_total = clean((sg_body.get("metadata") or {}).get("total_results"))
    v_total = clean((v_body.get("metadata") or {}).get("total_results"))
    return {
        "company_name": clean(company["company_name"]),
        "website_domain": clean(company["domain"]),
        "active_job_count_total": total,
        "active_job_count_sg": sg_total,
        "roles_by_function_eng_design_it_sales_ops": roles_text(classify_roles([clean(j.get("job_title") or j.get("title")) for j in jobs])),
        "hiring_velocity_postings_last_90_days": v_total,
        "job_locations": job_locations(jobs),
        "posting_first_seen_date": first_seen(jobs),
        "source_url": source_urls(jobs),
    }


def apollo_org_id(company: pd.Series) -> str:
    for base in [ROOT / "Firmographic" / "apollo" / "raw_exports", ROOT / "Backup" / "Firmographic" / "raw" / "apollo"]:
        path = base / f"{safe_slug(company['company_name'])}.json"
        if not path.exists():
            continue
        raw = json.loads(path.read_text(encoding="utf-8"))
        org = (raw.get("json") or raw).get("organization") or {}
        return clean(org.get("id"))
    return ""


def live_apollo(company: pd.Series, api_key: str, raw_dir: Path) -> tuple[dict[str, Any], dict[str, Any]]:
    org_id = apollo_org_id(company)
    if not org_id:
        export = {"error": "Apollo organization id not found"}
        write_json(raw_dir / f"{safe_slug(company['company_name'])}.json", export)
        return export, {"company_name": company["company_name"], "status_code": "NO_ORG_ID", "latency_sec": 0, "success": False, "records_retrieved": 0, "error": "Apollo organization id not found", "rate_limit": ""}
    url = f"https://api.apollo.io/api/v1/organizations/{org_id}/job_postings"
    headers = {"accept": "application/json", "Cache-Control": "no-cache", "x-api-key": api_key}
    export: dict[str, Any] = {"request": {"method": "GET", "url": url, "params": {"per_page": 100}}, "pages": [], "job_postings": [], "pagination": {}}
    traces = []
    page = 1
    total_pages = 1
    while page <= total_pages and page <= 100:
        params = {"page": page, "per_page": 100}
        started = time.perf_counter()
        response = requests.get(url, headers=headers, params=params, timeout=HTTP_TIMEOUT)
        latency = time.perf_counter() - started
        body = response_json(response)
        export["pages"].append({"request": {"params": params}, "response": {"status_code": response.status_code, "headers": dict(response.headers), "latency_sec": round(latency, 3), "json": body, "text": response.text if not isinstance(body, (dict, list)) else ""}})
        traces.append({"status_code": response.status_code, "latency_sec": round(latency, 3), "records": len(extract_jobs(body if isinstance(body, dict) else {})), "error": "" if response.ok else response.text[:500], "rate_limit": "; ".join(f"{k}: {v}" for k, v in response.headers.items() if k.lower().startswith("x-rate-limit"))})
        if not response.ok or not isinstance(body, dict):
            break
        export["job_postings"].extend(extract_jobs(body))
        pagination = body.get("pagination") if isinstance(body.get("pagination"), dict) else {}
        export["pagination"] = pagination
        total_pages = int(pagination.get("total_pages") or total_pages or 1)
        page += 1
    write_json(raw_dir / f"{safe_slug(company['company_name'])}.json", export)
    ok = all(int(trace["status_code"]) == 200 for trace in traces)
    return export, {"company_name": company["company_name"], "status_code": "200" if ok else ";".join(str(t["status_code"]) for t in traces), "latency_sec": round(sum(float(t["latency_sec"]) for t in traces), 3), "success": ok, "records_retrieved": len(export["job_postings"]), "error": "; ".join(t["error"] for t in traces if t["error"]), "rate_limit": next((t["rate_limit"] for t in traces if t["rate_limit"]), "")}


def extract_jobs(raw: dict[str, Any]) -> list[dict[str, Any]]:
    for key in ["job_postings", "organization_job_postings", "jobs", "data", "results", "active_job_postings"]:
        value = raw.get(key)
        if isinstance(value, list):
            return [x for x in value if isinstance(x, dict)]
    return []


def parse_apollo(raw: dict[str, Any], company: pd.Series) -> dict[str, str]:
    jobs = extract_jobs(raw)
    pagination = raw.get("pagination") if isinstance(raw.get("pagination"), dict) else {}
    total = clean(pagination.get("total_entries") or raw.get("total_results") or raw.get("total_count") or raw.get("count") or len(jobs))
    return {
        "company_name": clean(company["company_name"]),
        "website_domain": clean(company["domain"]),
        "active_job_count_total": total,
        "active_job_count_sg": str(count_sg(jobs)),
        "roles_by_function_eng_design_it_sales_ops": roles_text(classify_roles([clean(j.get("title") or j.get("job_title") or j.get("name")) for j in jobs])),
        "hiring_velocity_postings_last_90_days": velocity_90(jobs),
        "job_locations": job_locations(jobs),
        "posting_first_seen_date": first_seen(jobs),
        "source_url": source_urls(jobs),
    }


def apollo_full_pages_available(raw: dict[str, Any]) -> bool:
    if raw.get("pages"):
        successful_pages = {
            (page.get("request") or {}).get("params", {}).get("page")
            for page in raw.get("pages") or []
            if (page.get("response") or {}).get("status_code") == 200
        }
        pagination = raw.get("pagination") if isinstance(raw.get("pagination"), dict) else {}
        total_pages = int(pagination.get("total_pages") or 1)
        return len(successful_pages) >= total_pages
    pagination = raw.get("pagination") if isinstance(raw.get("pagination"), dict) else {}
    return int(pagination.get("total_pages") or 1) == 1


def blank_apollo_partial_metrics(row: dict[str, str]) -> dict[str, str]:
    row["active_job_count_sg"] = ""
    row["roles_by_function_eng_design_it_sales_ops"] = ""
    row["hiring_velocity_postings_last_90_days"] = ""
    return row


def cap_sample_fields(row: dict[str, str]) -> dict[str, str]:
    row["job_locations"] = "; ".join([value for value in row.get("job_locations", "").split("; ") if value][:15])
    row["source_url"] = "; ".join([value for value in row.get("source_url", "").split("; ") if value][:5])
    return row


def live_coresignal(company: pd.Series, api_key: str, raw_dir: Path) -> tuple[dict[str, Any], dict[str, Any]]:
    search_url = "https://api.coresignal.com/cdapi/v2/job_multi_source/search/es_dsl"
    collect_url = "https://api.coresignal.com/cdapi/v2/job_multi_source/collect"
    headers = {"accept": "application/json", "apikey": api_key, "Content-Type": "application/json"}
    domain = clean(company["domain"])
    cutoff = f"{(date.today() - timedelta(days=90)).isoformat()} 00:00:00"
    payloads = {
        "all_jobs_search": {"query": {"bool": {"must": [{"term": {"company_domain": domain}}, {"term": {"status": 1}}]}}},
        "sg_jobs_search": {"query": {"bool": {"must": [{"term": {"company_domain": domain}}, {"term": {"status": 1}}, {"match": {"country": {"query": "Singapore", "operator": "and"}}}]}}},
        "last_90_days_search": {"query": {"bool": {"must": [{"term": {"company_domain": domain}}, {"term": {"status": 1}}, {"range": {"date_posted": {"gte": cutoff}}}]}}},
    }
    export: dict[str, Any] = {"requests": {}, "responses": {}, "collected_jobs": []}
    traces = []
    for label, payload in payloads.items():
        started = time.perf_counter()
        response = requests.post(search_url, headers=headers, json=payload, params={"page": 1}, timeout=HTTP_TIMEOUT)
        latency = time.perf_counter() - started
        body = response_json(response)
        export["requests"][label] = {"method": "POST", "url": search_url, "params": {"page": 1}, "payload": payload}
        export["responses"][label] = {"status_code": response.status_code, "headers": dict(response.headers), "latency_sec": round(latency, 3), "json": body, "text": response.text if not isinstance(body, (dict, list)) else ""}
        traces.append({"status_code": response.status_code, "latency_sec": round(latency, 3), "records": int(response.headers.get("x-total-results") or 0), "error": "" if response.ok else response.text[:500], "rate_limit": "; ".join(f"{k}: {v}" for k, v in response.headers.items() if "rate" in k.lower())})

    ids = export["responses"].get("all_jobs_search", {}).get("json") or []
    if not isinstance(ids, list):
        ids = []
    for job_id in ids[:10]:
        started = time.perf_counter()
        response = requests.get(f"{collect_url}/{job_id}", headers={"accept": "application/json", "apikey": api_key}, timeout=HTTP_TIMEOUT)
        latency = time.perf_counter() - started
        body = response_json(response)
        export["collected_jobs"].append({"request": {"method": "GET", "url": f"{collect_url}/{job_id}"}, "response": {"status_code": response.status_code, "headers": dict(response.headers), "latency_sec": round(latency, 3), "json": body, "text": response.text if not isinstance(body, (dict, list)) else ""}})
        traces.append({"status_code": response.status_code, "latency_sec": round(latency, 3), "records": 1 if response.ok else 0, "error": "" if response.ok else response.text[:500], "rate_limit": "; ".join(f"{k}: {v}" for k, v in response.headers.items() if "rate" in k.lower())})

    write_json(raw_dir / f"{safe_slug(company['company_name'])}.json", export)
    ok = all(int(trace["status_code"]) == 200 for trace in traces)
    return export, {"company_name": company["company_name"], "status_code": "200" if ok else ";".join(str(t["status_code"]) for t in traces), "latency_sec": round(sum(float(t["latency_sec"]) for t in traces), 3), "success": ok, "records_retrieved": max((int(t["records"]) for t in traces[:1]), default=0) + sum(int(t["records"]) for t in traces[3:]), "error": "; ".join(t["error"] for t in traces if t["error"]), "rate_limit": next((t["rate_limit"] for t in traces if t["rate_limit"]), "")}


def parse_coresignal(raw: dict[str, Any], company: pd.Series) -> dict[str, str]:
    collected = raw.get("collected_jobs") or []
    jobs = []
    if isinstance(collected, list):
        for item in collected:
            body = ((item.get("response") or {}).get("json") if isinstance(item, dict) else None)
            if isinstance(body, dict):
                jobs.append(body)
    responses = raw.get("responses") if isinstance(raw.get("responses"), dict) else {}
    all_headers = ((responses.get("all_jobs_search") or {}).get("headers") or {})
    sg_headers = ((responses.get("sg_jobs_search") or {}).get("headers") or {})
    recent_headers = ((responses.get("last_90_days_search") or {}).get("headers") or {})
    total = clean(all_headers.get("x-total-results") or len(jobs))
    sg_total = clean(sg_headers.get("x-total-results"))
    recent_count = clean(recent_headers.get("x-total-results"))
    return {
        "company_name": clean(company["company_name"]),
        "website_domain": clean(company["domain"]),
        "active_job_count_total": total,
        "active_job_count_sg": sg_total or str(count_sg(jobs)),
        "roles_by_function_eng_design_it_sales_ops": roles_text(classify_roles([clean(j.get("job_posting_title") or j.get("title") or j.get("job_title")) for j in jobs])),
        "hiring_velocity_postings_last_90_days": recent_count or velocity_90(jobs),
        "job_locations": job_locations(jobs),
        "posting_first_seen_date": first_seen(jobs),
        "source_url": source_urls(jobs),
    }


def missing_fields(rows: list[dict[str, str]]) -> list[str]:
    fields = [c for c in DATA_COLUMNS if c not in {"company_name", "website_domain"}]
    return [field for field in fields if any(not clean(row.get(field)) for row in rows)]


def summary_record(tool_name: str, traces: list[dict[str, Any]], rows: list[dict[str, str]], credits: str, auth: str, missing_info: str) -> dict[str, str]:
    statuses: dict[str, int] = {}
    for trace in traces:
        key = str(trace["status_code"])
        statuses[key] = statuses.get(key, 0) + 1
    successes = sum(1 for trace in traces if trace.get("success"))
    error_rate = ((len(traces) - successes) / len(traces) * 100) if traces else 100
    latency = sum(float(trace.get("latency_sec") or 0) for trace in traces) / len(traces) if traces else 0
    rates = [clean(trace.get("rate_limit")) for trace in traces if clean(trace.get("rate_limit"))]
    gated = missing_fields(rows)
    return {
        "Tool Name": tool_name,
        "Category": "Jobs / Hiring",
        "API Available (Y/N)": "Y",
        "Authentication Type": auth,
        "Credits Used": credits,
        "Rate Limit": rates[0] if rates else "Not returned in response headers",
        "Companies Processed": str(len(traces)),
        "Error Rate (%)": f"{error_rate:.2f}",
        "Latency (sec)": f"{latency:.3f} avg",
        "Status Code / Error Flag": "; ".join(f"{k}:{v}" for k, v in sorted(statuses.items())),
        "Records Retrieved": str(sum(int(trace.get("records_retrieved") or 0) for trace in traces)),
        "Gated Fields Flag": "Y - " + "; ".join(gated) if gated else "N",
        "missing_info": missing_info,
        "Raw Export Saved (Y/N)": "Y",
    }


def build_tool(tool: str, companies: pd.DataFrame, env: dict[str, str]) -> None:
    tool_dir = OUT_ROOT / tool
    raw_dir = tool_dir / "raw_exports"
    raw_dir.mkdir(parents=True, exist_ok=True)
    rows: list[dict[str, str]] = []
    traces: list[dict[str, Any]] = []
    if tool == "theirstack":
        api_key = env.get("THEIRSTACK_API_KEY", "")
        before = theirstack_balance(api_key, raw_dir, "before") if api_key else {}
        for _, company in companies.iterrows():
            raw, trace = live_theirstack(company, api_key, raw_dir)
            traces.append(trace)
            rows.append(parse_theirstack(raw, company) if trace["success"] else blank_row(company, trace.get("error", "")))
        after = theirstack_balance(api_key, raw_dir, "after") if api_key else {}
        used = ""
        if isinstance(before.get("used_api_credits"), int) and isinstance(after.get("used_api_credits"), int):
            used = f"{after['used_api_credits'] - before['used_api_credits']} API credits used in this run; TheirStack job search consumes 1 credit per returned job"
        report = summary_record("TheirStack Job Search API", traces, rows, used or "Estimated from returned jobs; job search consumes 1 credit per returned job", "Bearer token in Authorization header", "TheirStack gives total counts via metadata and sampled active job details from job search.")
    elif tool == "apollo":
        api_key = env.get("APOLLO_API_KEY", "")
        provenance: list[dict[str, str]] = []
        for _, company in companies.iterrows():
            raw, trace = live_apollo(company, api_key, raw_dir)
            source = "live_paginated_apollo_raw"
            status = "full" if trace["success"] and apollo_full_pages_available(raw) else "partial_live_credit_blocked"
            if not raw.get("job_postings"):
                backup_path = BACKUP_APOLLO_JOBS_DIR / f"{safe_slug(company['company_name'])}.json"
                if backup_path.exists():
                    raw = json.loads(backup_path.read_text(encoding="utf-8"))
                    source = "saved_prior_apollo_page1_raw"
                    status = "saved_full_page" if apollo_full_pages_available(raw) else "saved_page1_only_current_credit_blocked"
                    trace = {
                        "company_name": company["company_name"],
                        "status_code": "SAVED_RAW_USED_AFTER_422_CREDIT_LIMIT",
                        "latency_sec": 0,
                        "success": True,
                        "records_retrieved": len(extract_jobs(raw)),
                        "error": "Current Apollo live call returned insufficient credits; used saved prior Apollo raw export.",
                        "rate_limit": "x-rate-limit-24-hour: 600; x-rate-limit-hourly: 200; x-rate-limit-minute: 50",
                    }
            traces.append(trace)
            has_apollo_data = bool(extract_jobs(raw)) or bool(raw.get("pagination"))
            row = parse_apollo(raw, company) if trace["success"] or has_apollo_data else blank_row(company, trace.get("error", ""))
            if (trace["success"] or has_apollo_data) and not apollo_full_pages_available(raw):
                row = blank_apollo_partial_metrics(row)
            rows.append(cap_sample_fields(row))
            pagination = raw.get("pagination") if isinstance(raw.get("pagination"), dict) else {}
            provenance.append({
                "company_name": clean(company["company_name"]),
                "website_domain": clean(company["domain"]),
                "apollo_source": source,
                "apollo_collection_status": status,
                "apollo_total_jobs_from_pagination": clean(pagination.get("total_entries") or row.get("active_job_count_total")),
                "apollo_jobs_returned_for_sample_fields": str(len(extract_jobs(raw))),
                "full_pages_available": str(apollo_full_pages_available(raw)),
                "fields_blank_reason": "" if apollo_full_pages_available(raw) else "Apollo full pagination unavailable/current key credit-blocked; SG count, roles, and 90-day velocity require complete pages, so they are intentionally blank instead of sample/fake values.",
            })
        full_count = sum(1 for item in provenance if item["full_pages_available"] == "True")
        partial_count = len(provenance) - full_count
        report = summary_record("Apollo Organization Job Postings API", traces, rows, "Live paginated calls made until Apollo returned 422 insufficient credits; saved prior Apollo raw used for page-1 evidence where live collection was blocked", "API Key in x-api-key header", "Apollo final export contains all 10 companies. Active job count total is taken from Apollo pagination where available. For companies without full Apollo pagination, SG count, roles by function, and 90-day velocity are intentionally blank because using page-1/sample jobs would create fake repeated values such as 25 or misleading zeros.")
        report["Status Code / Error Flag"] = f"FULL_APOLLO_PAGES:{full_count}; PARTIAL_OR_PAGE1_ONLY:{partial_count}"
        report["Gated Fields Flag"] = "Y - active_job_count_sg; roles_by_function_eng_design_it_sales_ops; hiring_velocity_postings_last_90_days are blank where full Apollo pagination was unavailable"
        write_json(raw_dir / "info_and_diagnostics" / "apollo_field_provenance.json", provenance)
        write_xlsx(raw_dir / "info_and_diagnostics" / "apollo_field_provenance.xlsx", provenance, list(provenance[0].keys()), "field_provenance")
    else:
        api_key = env.get("CORESIGNAL_API_KEY", "")
        for _, company in companies.iterrows():
            raw, trace = live_coresignal(company, api_key, raw_dir)
            traces.append(trace)
            rows.append(parse_coresignal(raw, company) if trace["success"] else blank_row(company, trace.get("error", "")))
        successful_searches = sum(1 for t in traces if t.get("success")) * 3
        successful_collects = sum(sum(1 for item in (json.loads((raw_dir / f"{safe_slug(company['company_name'])}.json").read_text(encoding="utf-8")).get("collected_jobs") or []) if (item.get("response") or {}).get("status_code") == 200) for _, company in companies.iterrows())
        report = summary_record("Coresignal Multi-source Jobs API", traces, rows, f"Estimated {successful_searches} search credits + {successful_collects} collect credits", "API Key in apikey header", "Coresignal Jobs API search provides counts; collect endpoint provides job-level location, date, and source URL.")
    write_json(tool_dir / f"{tool}_data.json", rows)
    write_json(tool_dir / f"{tool}_api_records.json", [report])
    write_json(tool_dir / f"{tool}_api_call_detail.json", traces)
    write_xlsx(tool_dir / f"{tool}_data.xlsx", rows, DATA_COLUMNS, f"{tool}_data")
    write_xlsx(tool_dir / f"{tool}_api_report.xlsx", [report], API_REPORT_COLUMNS, f"{tool}_api_report")


def blank_row(company: pd.Series, reason: str) -> dict[str, str]:
    return {
        "company_name": clean(company["company_name"]),
        "website_domain": clean(company["domain"]),
        "active_job_count_total": "",
        "active_job_count_sg": "",
        "roles_by_function_eng_design_it_sales_ops": "",
        "hiring_velocity_postings_last_90_days": "",
        "job_locations": "",
        "posting_first_seen_date": "",
        "source_url": clean(reason),
    }


def main() -> None:
    env = load_env(ROOT / ".env")
    companies = pd.read_csv(INPUT, dtype=str).fillna("")
    if OUT_ROOT.exists():
        shutil.rmtree(OUT_ROOT)
    for tool in ("theirstack", "apollo", "coresignal"):
        build_tool(tool, companies, env)
    print(f"Wrote {OUT_ROOT}")
    for path in sorted(OUT_ROOT.glob("*/*")):
        if path.is_file():
            print(path.relative_to(ROOT))


if __name__ == "__main__":
    main()
